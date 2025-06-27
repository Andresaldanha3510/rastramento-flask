import eventlet
eventlet.monkey_patch()

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from datetime import datetime, timedelta
import requests
import logging
import os
import math
import re
from dotenv import load_dotenv
import boto3
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy.exc import IntegrityError
from sqlalchemy import or_
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from flask import jsonify
from flask import make_response
from sqlalchemy import UniqueConstraint
from num2words import num2words
from collections import defaultdict
from flask_socketio import SocketIO, emit, join_room, leave_room



# ---- Configurações Iniciais ----
load_dotenv()

import os
print("ENDPOINT:", os.getenv('CLOUDFLARE_R2_ENDPOINT'))
print("ACCESS_KEY:", os.getenv('CLOUDFLARE_R2_ACCESS_KEY'))
print("SECRET_KEY:", os.getenv('CLOUDFLARE_R2_SECRET_KEY'))
print("BUCKET:", os.getenv('CLOUDFLARE_R2_BUCKET'))
print("PUBLIC_URL:", os.getenv('CLOUDFLARE_R2_PUBLIC_URL'))

app = Flask(__name__)
                                                                                                                

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'trackgo789@gmail.com'
app.config['MAIL_PASSWORD'] = 'mmoa moxc juli sfbe'
app.config['MAIL_DEFAULT_SENDER'] = 'trackgo789@gmail.com'

mail = Mail(app)

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///transport.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'w9z$kL2mNpQvR7tYxJ3hF8gWcPqB5vM2nZ4rT6yU')
Maps_API_KEY = os.getenv('Maps_API_KEY', 'AIzaSyBPdSOZF2maHURmdRmVzLgVo5YO2wliylo')
GEOAPIFY_API_KEY = os.getenv('GEOAPIFY_API_KEY', '7cd423ef184f48f0b770682cbebe11d0') # Usar os.getenv para Geoapify também
# app.config['OPENCAGE_API_KEY'] = os.getenv('OPENCAGE_API_KEY') # REMOVIDO: Não mais utilizado

app.config['CLOUDFLARE_R2_ENDPOINT'] = os.getenv('CLOUDFLARE_R2_ENDPOINT')
app.config['CLOUDFLARE_R2_ACCESS_KEY'] = os.getenv('CLOUDFLARE_R2_ACCESS_KEY')
app.config['CLOUDFLARE_R2_SECRET_KEY'] = os.getenv('CLOUDFLARE_R2_SECRET_KEY')
app.config['CLOUDFLARE_R2_BUCKET'] = os.getenv('CLOUDFLARE_R2_BUCKET')
app.config['CLOUDFLARE_R2_PUBLIC_URL'] = os.getenv('CLOUDFLARE_R2_PUBLIC_URL')


logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

db = SQLAlchemy(app)
migrate = Migrate(app, db)
socketio = SocketIO(app, async_mode='eventlet')

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    """Carrega o usuário pelo ID."""
    # Alterado para Session.get() como recomendado pelo SQLAlchemy 2.0
    return db.session.get(Usuario, int(user_id))


# ---- Modelos do Banco de Dados ----
class Motorista(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    data_nascimento = db.Column(db.Date, nullable=False)
    endereco = db.Column(db.String(200), nullable=False)
    pessoa_tipo = db.Column(db.String(10), nullable=False)
    cpf_cnpj = db.Column(db.String(14), unique=True, nullable=False, index=True)
    rg = db.Column(db.String(9), nullable=True)
    telefone = db.Column(db.String(11), nullable=False)
    cnh = db.Column(db.String(11), unique=True, nullable=False, index=True)
    validade_cnh = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    anexos = db.Column(db.String(500), nullable=True)
    
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    usuario = db.relationship('Usuario', backref='motorista', uselist=False)
    viagens = db.relationship('Viagem', backref='motorista_formal')

class Licenca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    
    # Chave estrangeira para a empresa (relação um-para-um)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False, unique=True)
    
    plano = db.Column(db.String(50), nullable=False, default='Básico') # Ex: Básico, Pro, Enterprise
    max_usuarios = db.Column(db.Integer, nullable=False, default=5)
    max_veiculos = db.Column(db.Integer, nullable=False, default=5)
    data_expiracao = db.Column(db.Date, nullable=True)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<Licenca {self.id} - Plano {self.plano} para Empresa {self.empresa_id}>'

 

import uuid
from datetime import datetime, timedelta

class Convite(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False, index=True)
    token = db.Column(db.String(36), unique=True, nullable=False, index=True)
    usado = db.Column(db.Boolean, default=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_expiracao = db.Column(db.DateTime, nullable=False)
    criado_por = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='Motorista')
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)

class Empresa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    razao_social = db.Column(db.String(200), nullable=False)
    nome_fantasia = db.Column(db.String(200), nullable=True)
    cnpj = db.Column(db.String(14), unique=True, nullable=False, index=True)
    inscricao_estadual = db.Column(db.String(20), nullable=True)
    endereco = db.Column(db.String(255), nullable=False)
    cidade = db.Column(db.String(100), nullable=False)
    estado = db.Column(db.String(2), nullable=False)
    cep = db.Column(db.String(8), nullable=False)
    telefone = db.Column(db.String(11), nullable=True)
    email_contato = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    licenca = db.relationship('Licenca', backref='empresa', uselist=False, cascade="all, delete-orphan")
    usuarios = db.relationship('Usuario', backref='empresa', lazy=True)

    def __repr__(self):
        return f'<Empresa {self.razao_social}>'
    
# Em app.py, adicione este novo modelo junto com os outros (Cliente, Viagem, etc.)

class Cobranca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False) # Quem gerou a cobrança

    valor_total = db.Column(db.Float, nullable=False)
    data_emissao = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    data_vencimento = db.Column(db.Date, nullable=False)
    data_pagamento = db.Column(db.DateTime, nullable=True)

    status = db.Column(db.String(20), nullable=False, default='Pendente', index=True) # Pendente, Paga, Vencida
    meio_pagamento = db.Column(db.String(50), nullable=True)
    observacoes = db.Column(db.Text, nullable=True)

    # Relacionamentos
    cliente = db.relationship('Cliente', backref='cobrancas')
    usuario = db.relationship('Usuario', backref='cobrancas_geradas')
    viagens = db.relationship('Viagem', backref='cobranca', lazy='dynamic')

    @property
    def is_vencida(self):
        """Verifica se a cobrança está vencida."""
        return self.data_vencimento < datetime.utcnow().date() and self.status == 'Pendente'

    def __repr__(self):
        return f'<Cobranca {self.id} - Cliente {self.cliente.nome_razao_social}>'

class Romaneio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Usaremos o ID como número do romaneio para simplificar
    data_emissao = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    observacoes = db.Column(db.Text, nullable=True)
    
    # Chave estrangeira para vincular o romaneio à viagem
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False, unique=True)
    viagem = db.relationship('Viagem', backref=db.backref('romaneio', uselist=False))
    
    # Relação com os itens da carga
    itens = db.relationship('ItemRomaneio', backref='romaneio', lazy=True, cascade="all, delete-orphan")

    @property
    def total_volumes(self):
        return len(self.itens) if self.itens else 0

    @property
    def peso_total(self):
        return sum(item.peso_total_item for item in self.itens) if self.itens else 0.0

class ItemRomaneio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    produto_descricao = db.Column(db.String(255), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False, default=1)
    embalagem = db.Column(db.String(50), nullable=True)
    peso_kg = db.Column(db.Float, nullable=True, default=0.0)
    
    # Chave estrangeira para vincular ao romaneio
    romaneio_id = db.Column(db.Integer, db.ForeignKey('romaneio.id'), nullable=False)

    @property
    def peso_total_item(self):
        return (self.peso_kg or 0.0) * (self.quantidade or 0)

class Veiculo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    placa = db.Column(db.String(7), unique=True, nullable=False, index=True)
    categoria = db.Column(db.String(50), nullable=True)
    modelo = db.Column(db.String(50), nullable=False)
    ano = db.Column(db.Integer, nullable=True)
    valor = db.Column(db.Float, nullable=True)
    km_rodados = db.Column(db.Float, nullable=True)
    ultima_manutencao = db.Column(db.Date, nullable=True)
    disponivel = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    viagens = db.relationship('Viagem', backref='veiculo')

from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import UserMixin

class Usuario(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    sobrenome = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    senha_hash = db.Column(db.String(128), nullable=False)
    telefone = db.Column(db.String(11), nullable=True)
    idioma = db.Column(db.String(20), default='Português')
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_phone = db.Column(db.String(11), nullable=True)
    is_admin = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(20), nullable=False, default='Motorista')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    cpf_cnpj = db.Column(db.String(14), unique=True, nullable=True, index=True)

    # CAMPO ADICIONADO: Chave estrangeira para a empresa.
    # Um usuário pertence a uma empresa.
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)

    def set_password(self, password):
        self.senha_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.senha_hash, password)
    
class Destino(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False)
    endereco = db.Column(db.String(200), nullable=False)
    ordem = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Abastecimento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=False)
    data_abastecimento = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    litros = db.Column(db.Float, nullable=False)
    preco_por_litro = db.Column(db.Float, nullable=False)
    custo_total = db.Column(db.Float, nullable=False)
    odometro = db.Column(db.Float, nullable=False)
    anexo_comprovante = db.Column(db.String(500), nullable=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False, index=True) 
    data_abastecimento = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    veiculo = db.relationship('Veiculo', backref='abastecimentos')
    motorista = db.relationship('Motorista', backref='abastecimentos_registrados')

class CustoViagem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=False, unique=True)
    combustivel = db.Column(db.Float, nullable=True)
    pedagios = db.Column(db.Float, nullable=True)
    alimentacao = db.Column(db.Float, nullable=True)
    hospedagem = db.Column(db.Float, nullable=True)
    outros = db.Column(db.Float, nullable=True)
    descricao_outros = db.Column(db.String(300), nullable=True)
    anexos = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    viagem = db.relationship('Viagem', backref=db.backref('custo_viagem', uselist=False))

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    pessoa_tipo = db.Column(db.String(10), nullable=False)  # 'fisica' ou 'juridica'
    nome_razao_social = db.Column(db.String(200), nullable=False)
    nome_fantasia = db.Column(db.String(200), nullable=True)
    cpf_cnpj = db.Column(db.String(14), unique=True, nullable=False, index=True)
    inscricao_estadual = db.Column(db.String(20), nullable=True)

    cep = db.Column(db.String(8), nullable=False)
    logradouro = db.Column(db.String(255), nullable=False)
    numero = db.Column(db.String(20), nullable=False)
    complemento = db.Column(db.String(100), nullable=True)
    bairro = db.Column(db.String(100), nullable=False)
    cidade = db.Column(db.String(100), nullable=False)
    estado = db.Column(db.String(2), nullable=False)

    email = db.Column(db.String(120), nullable=False)
    telefone = db.Column(db.String(11), nullable=False)
    contato_principal = db.Column(db.String(100), nullable=True)

  
    anexos = db.Column(db.String(1000), nullable=True)  
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    
    cadastrado_por_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    cadastrado_por = db.relationship('Usuario', backref='clientes_cadastrados')

    def __repr__(self):
        return f'<Cliente {self.id}: {self.nome_razao_social}>'

class Viagem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=True)
    
    motorista_cpf_cnpj = db.Column(db.String(14), nullable=True, index=True)

    veiculo_id = db.Column(db.Integer, db.ForeignKey('veiculo.id'), nullable=False)
    cliente = db.Column(db.String(100), nullable=False)
    valor_recebido = db.Column(db.Float, nullable=True)
    forma_pagamento = db.Column(db.String(50), nullable=True)
    endereco_saida = db.Column(db.String(200), nullable=False)
    endereco_destino = db.Column(db.String(200), nullable=False)
    distancia_km = db.Column(db.Float, nullable=True)
    data_inicio = db.Column(db.DateTime, nullable=False)
    data_fim = db.Column(db.DateTime, nullable=True)
    duracao_segundos = db.Column(db.Integer, nullable=True)
    custo = db.Column(db.Float, nullable=True)
    status = db.Column(db.String(50), nullable=False, default='pendente', index=True)
    observacoes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    destinos = db.relationship('Destino', backref='viagem', lazy='dynamic', cascade='all, delete-orphan')
    cobranca_id = db.Column(db.Integer, db.ForeignKey('cobranca.id'), nullable=True)
    abastecimentos = db.relationship('Abastecimento', backref='viagem', lazy=True)

class Localizacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    motorista_id = db.Column(db.Integer, db.ForeignKey('motorista.id'), nullable=False)
    viagem_id = db.Column(db.Integer, db.ForeignKey('viagem.id'), nullable=True)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    endereco = db.Column(db.String(200), nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

from functools import wraps

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Acesso restrito ao administrador.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def owner_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'Owner':
            flash('Acesso restrito ao proprietário do sistema.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or (current_user.role not in ['Admin', 'Master']):
            flash('Acesso restrito a administradores ou masters.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def validate_cpf_cnpj(cpf_cnpj, pessoa_tipo):
    if pessoa_tipo == 'fisica':
        return bool(re.match(r'^\d{11}$', cpf_cnpj))
    return bool(re.match(r'^\d{14}$', cpf_cnpj))

def validate_telefone(telefone):
    return bool(re.match(r'^\d{10,11}$', telefone))

def validate_cnh(cnh):
    return bool(re.match(r'^\d{11}$', cnh))

def validate_placa(placa):
    return bool(re.match(r'^[A-Z0-9]{7}$', placa.upper()))

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

@app.route('/cadastrar_cliente', methods=['GET', 'POST'])
@login_required
def cadastrar_cliente():
    if request.method == 'POST':
        try:
            # 1. Coletar dados do formulário
            pessoa_tipo = request.form.get('pessoa_tipo')
            nome_razao_social = request.form.get('nome_razao_social')
            nome_fantasia = request.form.get('nome_fantasia') if pessoa_tipo == 'juridica' else None
            # Remove caracteres não numéricos do CPF/CNPJ, telefone e CEP
            cpf_cnpj = re.sub(r'\D', '', request.form.get('cpf_cnpj', ''))
            inscricao_estadual = request.form.get('inscricao_estadual') if pessoa_tipo == 'juridica' else None
            cep = re.sub(r'\D', '', request.form.get('cep', ''))
            logradouro = request.form.get('logradouro')
            numero = request.form.get('numero')
            complemento = request.form.get('complemento')
            bairro = request.form.get('bairro')
            cidade = request.form.get('cidade')
            estado = request.form.get('estado')
            email = request.form.get('email')
            telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            contato_principal = request.form.get('contato_principal')

            # 2. Validação dos dados
            if not all([pessoa_tipo, nome_razao_social, cpf_cnpj, cep, logradouro, numero, bairro, cidade, estado, email, telefone]):
                flash('Erro: Todos os campos principais são obrigatórios.', 'error')
                return redirect(url_for('cadastrar_cliente'))

            if Cliente.query.filter_by(cpf_cnpj=cpf_cnpj).first():
                flash('Erro: Este CPF/CNPJ já está cadastrado.', 'error')
                return redirect(url_for('cadastrar_cliente'))

            # 3. Lógica para upload de anexos (similar ao cadastro de motorista)
            anexos_urls = []
            files = request.files.getlist('anexos')
            if files and any(f.filename for f in files):
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        # Salva em uma pasta específica para clientes, organizada por CPF/CNPJ
                        s3_path = f"clientes/{cpf_cnpj}/{filename}"
                        s3_client.upload_fileobj(file, bucket_name, s3_path, ExtraArgs={'ContentType': file.content_type})
                        public_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
                        anexos_urls.append(public_url)

            # 4. Criar e salvar o novo cliente no banco
            novo_cliente = Cliente(
                pessoa_tipo=pessoa_tipo,
                nome_razao_social=nome_razao_social,
                nome_fantasia=nome_fantasia,
                cpf_cnpj=cpf_cnpj,
                inscricao_estadual=inscricao_estadual,
                cep=cep,
                logradouro=logradouro,
                numero=numero,
                complemento=complemento,
                bairro=bairro,
                cidade=cidade,
                estado=estado,
                email=email,
                telefone=telefone,
                contato_principal=contato_principal,
                anexos=','.join(anexos_urls) if anexos_urls else None,
                cadastrado_por_id=current_user.id  # Associa o cliente ao usuário logado
            )

            db.session.add(novo_cliente)
            db.session.commit()
            
            flash('Cliente cadastrado com sucesso!', 'success')
            return redirect(url_for('consultar_clientes')) # Redireciona para a nova página de consulta

        except IntegrityError:
            db.session.rollback()
            flash('Erro de integridade de dados. O CPF/CNPJ provavelmente já existe.', 'error')
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao cadastrar cliente: {e}", exc_info=True)
            flash(f'Ocorreu um erro inesperado: {e}', 'error')

    # Se a requisição for GET, apenas renderiza a página
    return render_template('cadastrar_cliente.html', active_page='cadastrar_cliente')




from sqlalchemy import or_ 
@app.route('/api/clientes/search')
@login_required
def search_clientes():
    search_term = request.args.get('term', '')
    if not search_term or len(search_term) < 2:
        return jsonify([])

  
    clientes = Cliente.query.filter(
        or_(
            Cliente.nome_razao_social.ilike(f'%{search_term}%'),
            Cliente.nome_fantasia.ilike(f'%{search_term}%')
        )
    ).limit(10).all()

   
    results = []
    search_term_lower = search_term.lower()

    for cliente in clientes:
        
        if cliente.nome_razao_social and search_term_lower in cliente.nome_razao_social.lower():
            results.append(cliente.nome_razao_social)
        
       
        if cliente.nome_fantasia and search_term_lower in cliente.nome_fantasia.lower() and cliente.nome_fantasia not in results:
            results.append(cliente.nome_fantasia)
    
    return jsonify(results)

@app.route('/registrar_abastecimento', methods=['POST'])
@login_required
def registrar_abastecimento():
    try:
        motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
        if not motorista_formal:
             return jsonify({'success': False, 'message': 'Perfil de motorista formal não encontrado para o usuário atual.'}), 400

        viagem_ativa = Viagem.query.filter(
            or_(
                Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj,
                Viagem.motorista_id == motorista_formal.id
            ),
            Viagem.status == 'em_andamento'
        ).first()

        if not viagem_ativa:
            return jsonify({'success': False, 'message': 'Nenhuma viagem ativa encontrada para associar o abastecimento.'}), 400

        litros = float(request.form.get('litros'))
        preco_por_litro = float(request.form.get('preco_por_litro'))
        odometro = float(request.form.get('odometro'))
        custo_total = litros * preco_por_litro

        novo_abastecimento = Abastecimento(
            veiculo_id=viagem_ativa.veiculo_id,
            motorista_id=motorista_formal.id,
            viagem_id=viagem_ativa.id,
            litros=litros,
            preco_por_litro=preco_por_litro,
            custo_total=custo_total,
            odometro=odometro
        )

        anexo_url = None
        anexo = request.files.get('anexo_comprovante')
        if anexo and anexo.filename:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            filename = secure_filename(anexo.filename)
            s3_path = f"abastecimentos/{viagem_ativa.id}/{uuid.uuid4()}-{filename}"
            
            s3_client.upload_fileobj(
                anexo,
                bucket_name,
                s3_path,
                ExtraArgs={'ContentType': anexo.content_type or 'application/octet-stream'}
            )
            anexo_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
        
        novo_abastecimento.anexo_comprovante = anexo_url

        db.session.add(novo_abastecimento)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Abastecimento registrado com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao registrar abastecimento: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Erro interno: {e}'}), 500
    

    

@app.route('/consultar_clientes')
@login_required
def consultar_clientes():
    search_query = request.args.get('search', '').strip()
    query = Cliente.query

    if search_query:
        search_filter = f"%{search_query}%"
        # Filtra por nome, cpf/cnpj ou cidade
        query = query.filter(
            or_(
                Cliente.nome_razao_social.ilike(search_filter),
                Cliente.cpf_cnpj.ilike(search_filter),
                Cliente.cidade.ilike(search_filter)
            )
        )
    
    # Ordena os clientes por nome
    clientes = query.order_by(Cliente.nome_razao_social.asc()).all()
    
    return render_template('consultar_clientes.html', 
                           clientes=clientes, 
                           search_query=search_query, 
                           active_page='consultar_clientes')

@app.route('/registrar/<token>', methods=['GET', 'POST'])
def registrar_com_token(token):
    convite = Convite.query.filter_by(token=token, usado=False).first()

    if not convite or convite.data_expiracao < datetime.utcnow():
        flash('Convite inválido ou expirado.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = request.form.get('nome')
        sobrenome = request.form.get('sobrenome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        cpf_cnpj = request.form.get('cpf_cnpj')

        if not validate_cpf_cnpj(cpf_cnpj, 'fisica') and not validate_cpf_cnpj(cpf_cnpj, 'juridica'):
             flash('CPF/CNPJ inválido. Deve conter 11 ou 14 dígitos numéricos.', 'error')
             return redirect(request.url)

        if email != convite.email:
            flash('E-mail diferente do convite.', 'error')
            return redirect(request.url)

        if Usuario.query.filter_by(email=email).first():
            flash('E-mail já cadastrado.', 'error')
            return redirect(request.url)

        if Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first():
            flash('CPF/CNPJ já cadastrado para outro usuário.', 'error')
            return redirect(request.url)

        # LÓGICA ATUALIZADA: Associa o novo usuário à empresa vinda do convite
        usuario = Usuario(
            nome=nome,
            sobrenome=sobrenome,
            email=email,
            role=convite.role,
            is_admin=convite.role == 'Admin',
            cpf_cnpj=cpf_cnpj,
            empresa_id=convite.empresa_id
        )
        usuario.set_password(senha)
        db.session.add(usuario)
        db.session.flush()

        if convite.role == 'Motorista':
            motorista_existente = Motorista.query.filter_by(cpf_cnpj=cpf_cnpj).first()
            if not motorista_existente:
                novo_motorista_formal = Motorista(
                    nome=f"{usuario.nome} {usuario.sobrenome}",
                    data_nascimento=datetime(1900, 1, 1).date(),
                    endereco="A ser preenchido",
                    pessoa_tipo="fisica",
                    cpf_cnpj=cpf_cnpj,
                    rg=None,
                    telefone=usuario.telefone or "00000000000",
                    cnh=f"0000000000{str(uuid.uuid4().hex[:5])}",
                    validade_cnh=datetime.utcnow().date() + timedelta(days=365*5),
                    usuario_id=usuario.id
                )
                db.session.add(novo_motorista_formal)
            else:
                motorista_existente.usuario_id = usuario.id
                db.session.add(motorista_existente)

        convite.usado = True
        db.session.commit()
        flash('Conta criada com sucesso!', 'success')
        return redirect(url_for('login'))

    return render_template('registrar_token.html', email=convite.email, role=convite.role)


def get_coordinates(endereco):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {'address': endereco, 'key': Maps_API_KEY}
    try:
        logger.debug(f"Obtendo coordenadas para: {endereco}")
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        if data['status'] == 'OK' and data['results']:
            location = data['results'][0]['geometry']['location']
            return location['lat'], location['lng']
        logger.warning(f"Endereço não encontrado: {endereco}")
        return None, None
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro ao obter coordenadas: {str(e)}")
        return None, None

@app.route('/enviar_convite', methods=['POST'])
@login_required
@master_required
def enviar_convite():
    empresa_admin = current_user.empresa
    if empresa_admin and empresa_admin.licenca:
        usuarios_atuais = len(empresa_admin.usuarios)
        max_permitido = empresa_admin.licenca.max_usuarios
        if usuarios_atuais >= max_permitido:
            flash(f'Não é possível enviar o convite. Sua empresa atingiu o limite de {max_permitido} usuários do seu plano.', 'error')
            return redirect(url_for('configuracoes'))

    email = request.form.get('email')
    role = request.form.get('role')
    
    if not email or not role:
        flash('E-mail e papel são obrigatórios.', 'error')
        return redirect(url_for('configuracoes'))
    
    if current_user.role == 'Master' and role != 'Motorista':
        flash('Usuários do tipo Master podem convidar apenas Motoristas.', 'error')
        return redirect(url_for('configuracoes'))
    
    if role not in ['Motorista', 'Master']:
        flash('Papel inválido.', 'error')
        return redirect(url_for('configuracoes'))

    token = str(uuid.uuid4())
    data_expiracao = datetime.utcnow() + timedelta(days=3)

    convite = Convite(
        email=email, 
        token=token, 
        criado_por=current_user.id, 
        data_expiracao=data_expiracao, 
        role=role,
        empresa_id=current_user.empresa_id
    )
    db.session.add(convite)
    db.session.commit()

    link_convite = url_for('registrar_com_token', token=token, _external=True)
    msg = Message(
        subject=f'Convite para acessar o sistema como {role}',
        recipients=[email],
        body=f'Você foi convidado a se registrar no sistema como {role}. Clique no link abaixo:\n{link_convite}'
    )
    try:
        mail.send(msg)
        flash(f'Convite enviado para {email} como {role}!', 'success')
    except Exception as e:
        flash(f'Erro ao enviar o e-mail: {str(e)}', 'error')

    return redirect(url_for('configuracoes'))

def validar_endereco(endereco):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {'address': endereco, 'key': Maps_API_KEY}
    try:
        logger.debug(f"Validando endereço: {endereco}")
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        return data['status'] == 'OK' and len(data['results']) > 0
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro na validação de endereço: {str(e)}")
        return False


def calcular_distancia_e_duracao(enderecos):
    if len(enderecos) < 2:
        logger.error("Pelo menos dois endereços são necessários (origem e pelo menos um destino).")
        return None, None

    url = 'https://maps.googleapis.com/maps/api/directions/json'
    origem = enderecos[0]
    destinos = enderecos[1:]
    total_distancia_km = 0
    total_duracao_segundos = 0

    try:
        logger.debug(f"Calculando rota para endereços: {enderecos}")
        for i in range(len(destinos)):
            params = {
                'origin': origem,
                'destination': destinos[i],
                'key': Maps_API_KEY,
                'units': 'metric',
                'departure_time': 'now'
            }
            if i < len(destinos) - 1:
                params['waypoints'] = destinos[i]
                origem = destinos[i]
            response = requests.get(url, params=params, timeout=5)
            response.raise_for_status()
            data = response.json()
            if data['status'] == 'OK' and data['routes']:
                route = data['routes'][0]['legs'][0]
                total_distancia_km += route['distance']['value'] / 1000
                total_duracao_segundos += route['duration']['value']
            else:
                logger.warning(f"Erro na Directions API para trecho {origem} -> {destinos[i]}: {data.get('error_message', 'Erro desconhecido')}")
                lat1, lon1 = get_coordinates(origem)
                lat2, lon2 = get_coordinates(destinos[i])
                if lat1 is None or lat2 is None:
                    flash(f'Não foi possível calcular a distância para o trecho {origem} -> {destinos[i]}.', 'error')
                    return None, None
                distancia_km = haversine_distance(lat1, lon1, lat2, lon2)
                velocidade_media_kmh = 60
                duracao_segundos = int((distancia_km / velocidade_media_kmh) * 3600)
                total_distancia_km += distancia_km
                total_duracao_segundos += duracao_segundos
                flash(f'Distância calculada em linha reta para o trecho {origem} -> {destinos[i]}.', 'warning')
                origem = destinos[i]

        return total_distancia_km, total_duracao_segundos
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro na Directions API: {str(e)}")
        flash(f'Erro de conexão com a API do Google Maps: {str(e)}', 'error')
        return None, None

@app.route('/cobrancas')
@login_required
def consultar_cobrancas():
    # Query para buscar todas as cobranças, trazendo o cliente junto para evitar N+1 queries
    cobrancas = Cobranca.query.options(db.joinedload(Cobranca.cliente)).order_by(Cobranca.data_vencimento.desc()).all()
    
    # Atualiza o status para 'Vencida' se necessário
    for cobranca in cobrancas:
        if cobranca.is_vencida:
            cobranca.status = 'Vencida'
    db.session.commit()

    total_pendente = sum(c.valor_total for c in cobrancas if c.status in ['Pendente', 'Vencida'])
    total_pago = sum(c.valor_total for c in cobrancas if c.status == 'Paga')

    return render_template('consultar_cobrancas.html', 
                           cobrancas=cobrancas,
                           total_pendente=total_pendente,
                           total_pago=total_pago,
                           active_page='cobrancas')


@app.route('/cobranca/gerar', methods=['GET', 'POST'])
@login_required
def gerar_cobranca():
    if request.method == 'POST':
        try:
            cliente_id = request.form.get('cliente_id')
            viagem_ids = request.form.getlist('viagem_ids')
            data_vencimento_str = request.form.get('data_vencimento')
            observacoes = request.form.get('observacoes')

            if not all([cliente_id, viagem_ids, data_vencimento_str]):
                flash('Cliente, data de vencimento e ao menos uma viagem são obrigatórios.', 'error')
                return redirect(url_for('gerar_cobranca'))

            viagens_selecionadas = Viagem.query.filter(Viagem.id.in_(viagem_ids)).all()
            valor_total = sum(v.valor_recebido or 0 for v in viagens_selecionadas)

            nova_cobranca = Cobranca(
                cliente_id=cliente_id,
                usuario_id=current_user.id,
                valor_total=valor_total,
                data_vencimento=datetime.strptime(data_vencimento_str, '%Y-%m-%d').date(),
                observacoes=observacoes
            )
            
            db.session.add(nova_cobranca)
            
            for viagem in viagens_selecionadas:
                viagem.cobranca_id = nova_cobranca.id
            
            db.session.commit()
            
            flash('Cobrança gerada com sucesso! Visualizando a Nota de Débito.', 'success')
            return redirect(url_for('visualizar_nota_debito', cobranca_id=nova_cobranca.id))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao gerar cobrança: {e}", exc_info=True)
            flash(f'Ocorreu um erro ao gerar a cobrança: {e}', 'error')
    
    clientes = Cliente.query.order_by(Cliente.nome_razao_social).all()
    return render_template('gerar_cobranca.html', clientes=clientes, active_page='cobrancas')

@app.route('/api/cobranca/<int:cobranca_id>/marcar_paga', methods=['POST'])
@login_required
def api_marcar_paga(cobranca_id):
    cobranca = Cobranca.query.get_or_404(cobranca_id)
    try:
        data = request.get_json()
        meio_pagamento = data.get('meio_pagamento')

        if not meio_pagamento:
            return jsonify({'success': False, 'message': 'Meio de pagamento é obrigatório.'}), 400

        cobranca.status = 'Paga'
        cobranca.data_pagamento = datetime.utcnow()
        cobranca.meio_pagamento = meio_pagamento
        db.session.commit()

        return jsonify({'success': True, 'message': 'Cobrança marcada como paga.'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao marcar cobrança como paga: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


def get_address_geoapify(lat, lon):
    try:
        url = f'https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey={GEOAPIFY_API_KEY}'
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            if data['features']:
                return data['features'][0]['properties']['formatted']
    except Exception as e:
        logger.error(f"Erro na geocodificação Geoapify: {str(e)}")
    return "Endereço não encontrado"

@app.route('/api/cliente/<int:cliente_id>/viagens_nao_cobradas')
@login_required
def api_viagens_nao_cobradas(cliente_id):
    viagens = Viagem.query.filter_by(
        cliente=Cliente.query.get(cliente_id).nome_razao_social, 
        cobranca_id=None,
        status='concluida' # Apenas viagens concluídas podem ser cobradas
    ).all()
    
    viagens_data = [{
        'id': v.id,
        'data_inicio': v.data_inicio.strftime('%d/%m/%Y'),
        'destino': v.endereco_destino,
        'valor': v.valor_recebido or 0.0,
        'valor_formatado': f"R$ {v.valor_recebido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    } for v in viagens]
    
    return jsonify(viagens_data)


@app.route('/')
def index():
    motoristas = Motorista.query.all()
    veiculos = Veiculo.query.all()
    viagens_query = Viagem.query.all()
    
    viagens = []
    for viagem in viagens_query:
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        viagens.append({
            'id': viagem.id,
            'cliente': viagem.cliente,
            'motorista_nome': motorista_nome,
            'endereco_saida': viagem.endereco_saida,
            'endereco_destino': viagem.endereco_destino,
            'status': viagem.status,
            'veiculo_placa': viagem.veiculo.placa,
            'veiculo_modelo': viagem.veiculo.modelo,
            'data_inicio': viagem.data_inicio,
            'data_fim': viagem.data_fim,
            'distancia_km': viagem.distancia_km
        })

    return render_template(
        'index.html',
        motoristas=motoristas,
        veiculos=veiculos,
        viagens=viagens,
        Maps_API_KEY=Maps_API_KEY
    )


@app.route('/cadastrar_motorista', methods=['GET', 'POST'])
def cadastrar_motorista():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        data_nascimento = request.form.get('data_nascimento', '').strip()
        endereco = request.form.get('endereco', '').strip()
        pessoa_tipo = request.form.get('pessoa_tipo', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        rg = request.form.get('rg', '').strip() or None
        telefone = request.form.get('telefone', '').strip()
        cnh = request.form.get('cnh', '').strip()
        validade_cnh = request.form.get('validade_cnh', '').strip()
        files = request.files.getlist('anexos')

        if not all([nome, data_nascimento, endereco, pessoa_tipo, cpf_cnpj, telefone, cnh, validade_cnh]):
            flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
            return redirect(url_for('cadastrar_motorista'))

        try:
            data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
            validade_cnh = datetime.strptime(validade_cnh, '%Y-%m-%d').date()
        except ValueError:
            flash('Formato de data inválido.', 'error')
            return redirect(url_for('cadastrar_motorista'))

        if not validate_cpf_cnpj(cpf_cnpj, pessoa_tipo):
            flash(f"{'CPF' if pessoa_tipo == 'fisica' else 'CNPJ'} inválido. Deve conter {'11' if pessoa_tipo == 'fisica' else '14'} dígitos numéricos.", 'error')
            return redirect(url_for('cadastrar_motorista'))

        if not validate_telefone(telefone):
            flash('Telefone inválido. Deve conter 10 ou 11 dígitos numéricos.', 'error')
            return redirect(url_for('cadastrar_motorista'))

        if not validate_cnh(cnh):
            flash('CNH inválida. Deve conter 11 dígitos numéricos.', 'error')
            return redirect(url_for('cadastrar_motorista'))

        if Motorista.query.filter_by(cpf_cnpj=cpf_cnpj).first():
            flash('CPF/CNPJ já cadastrado.', 'error')
            return redirect(url_for('cadastrar_motorista'))
        if Motorista.query.filter_by(cnh=cnh).first():
            flash('CNH já cadastrada.', 'error')
            return redirect(url_for('cadastrar_motorista'))

        anexos_urls = []
        allowed_extensions = {'.pdf', '.jpg', '.jpeg', '.png'}
        if files and any(f.filename for f in files):
            try:
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']

                for file in files:
                    if file and file.filename:
                        extension = os.path.splitext(file.filename)[1].lower()
                        if extension not in allowed_extensions:
                            flash(f'Arquivo {file.filename} não permitido. Use PDF, JPG ou PNG.', 'error')
                            continue

                        filename = secure_filename(file.filename)
                        s3_path = f"motoristas/{cpf_cnpj}/{filename}"

                        s3_client.upload_fileobj(
                            file,
                            bucket_name,
                            s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        public_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
                        anexos_urls.append(public_url)
            except Exception as e:
                flash(f'Erro ao fazer upload dos arquivos: {str(e)}', 'error')
                return redirect(url_for('cadastrar_motorista'))

        motorista = Motorista(
            nome=nome,
            data_nascimento=data_nascimento,
            endereco=endereco,
            pessoa_tipo=pessoa_tipo,
            cpf_cnpj=cpf_cnpj,
            rg=rg,
            telefone=telefone,
            cnh=cnh,
            validade_cnh=validade_cnh,
            anexos=','.join(anexos_urls) if anexos_urls else None
        )

        try:
            db.session.add(motorista)
            db.session.commit()
            flash('Motorista cadastrado com sucesso!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar motorista: {str(e)}', 'error')
            return redirect(url_for('cadastrar_motorista'))

    return render_template('cadastrar_motorista.html', active_page='cadastrar_motorista')


@app.route('/editar_cliente/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    cliente = Cliente.query.get_or_404(cliente_id)

    if request.method == 'POST':
        try:
            novo_cpf_cnpj = re.sub(r'\D', '', request.form.get('cpf_cnpj', ''))

            cliente_existente = Cliente.query.filter(Cliente.cpf_cnpj == novo_cpf_cnpj, Cliente.id != cliente_id).first()
            if cliente_existente:
                flash('Erro: O CPF/CNPJ informado já pertence a outro cliente.', 'error')
                return redirect(url_for('editar_cliente', cliente_id=cliente_id))

            cliente.pessoa_tipo = request.form.get('pessoa_tipo')
            cliente.nome_razao_social = request.form.get('nome_razao_social')
            cliente.nome_fantasia = request.form.get('nome_fantasia') if cliente.pessoa_tipo == 'juridica' else None
            cliente.cpf_cnpj = novo_cpf_cnpj
            cliente.inscricao_estadual = request.form.get('inscricao_estadual') if cliente.pessoa_tipo == 'juridica' else None
            cliente.cep = re.sub(r'\D', '', request.form.get('cep', ''))
            cliente.logradouro = request.form.get('logradouro')
            cliente.numero = request.form.get('numero')
            cliente.complemento = request.form.get('complemento')
            cliente.bairro = request.form.get('bairro')
            cliente.cidade = request.form.get('cidade')
            cliente.estado = request.form.get('estado')
            cliente.email = request.form.get('email')
            cliente.telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            cliente.contato_principal = request.form.get('contato_principal')

            db.session.commit()
            flash('Cliente atualizado com sucesso!', 'success')
            return redirect(url_for('consultar_clientes'))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao editar cliente: {e}", exc_info=True)
            flash(f'Ocorreu um erro inesperado ao salvar as alterações: {e}', 'error')
            return redirect(url_for('editar_cliente', cliente_id=cliente_id))

    return render_template('editar_cliente.html', cliente=cliente, active_page='consultar_clientes')

@app.route('/nota_debito/<int:cobranca_id>')
@login_required
def visualizar_nota_debito(cobranca_id):
    """
    Exibe uma Nota de Débito específica com todos os dados para visualização e impressão.
    """
    cobranca = db.session.get(Cobranca, cobranca_id)
    if not cobranca:
        flash('Cobrança não encontrada.', 'error')
        return redirect(url_for('consultar_cobrancas'))


    empresa_emissora = None
    if cobranca.usuario and cobranca.usuario.empresa_id:
        empresa_emissora = db.session.get(Empresa, cobranca.usuario.empresa_id)


    valor_por_extenso = num2words(cobranca.valor_total, lang='pt_BR', to='currency')

    return render_template(
        'nota_debito.html',
        cobranca=cobranca,
        cliente=cobranca.cliente,
        viagens=cobranca.viagens.all(),
        empresa=empresa_emissora,
        valor_extenso=valor_por_extenso
    )

@app.route('/criar_admin')
def criar_admin():
    if not Usuario.query.filter_by(email='adminadmin@admin.com').first():
        admin = Usuario(
            nome='Admin',
            sobrenome='Master',
            email='adminadmin@admin.com'
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        return "Admin criado com sucesso"
    return "Admin já existe"


@app.route('/consultar_motoristas', methods=['GET'])
def consultar_motoristas():
    search_query = request.args.get('search', '').strip()
    query = Motorista.query
    if search_query:
        query = query.filter(
            (Motorista.nome.ilike(f'%{search_query}%')) |
            (Motorista.cpf_cnpj.ilike(f'%{search_query}%'))
        )
    motoristas = query.order_by(Motorista.nome.asc()).all()
    return render_template('consultar_motoristas.html', motoristas=motoristas, search_query=search_query, active_page='consultar_motoristas')

@app.route('/api/motorista/<int:motorista_id>/details')
@login_required # Protege a rota, garantindo que apenas usuários logados possam acessá-la.
def motorista_details_api(motorista_id):
    """Retorna os detalhes (estatísticas e viagens) de um motorista em formato JSON."""
    

    motorista = Motorista.query.get_or_404(motorista_id)
    
    viagens = motorista.viagens
    
    total_receita = sum(v.valor_recebido or 0 for v in viagens)
    # Supondo que você tenha um campo 'custo' no seu modelo Viagem.
    # Se não tiver, ajuste ou remova esta linha.
    total_custo = sum(getattr(v, 'custo', 0) or 0 for v in viagens)
    total_distancia = sum(getattr(v, 'distancia_km', 0) or 0 for v in viagens)

    stats = {
        'total_viagens': len(viagens),
        'total_receita': total_receita,
        'total_custo': total_custo,
        'lucro_total': total_receita - total_custo,
        'total_distancia': total_distancia
    }

    viagens_data = []
    for v in viagens:
        viagens_data.append({
            'cliente': v.cliente,
            'data_inicio': v.data_inicio.isoformat() if v.data_inicio else None,
            'endereco_saida': v.endereco_saida,
            'endereco_destino': v.endereco_destino,
            'status': v.status
        })

    return jsonify({
        'stats': stats,
        'viagens': viagens_data
    })



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        if not email or not senha:
            flash('Preencha todos os campos', 'error')
            return redirect(url_for('login'))
            
        usuario = Usuario.query.filter_by(email=email).first()
        
        if not usuario:
            flash('Usuário não encontrado', 'error')
            return redirect(url_for('login'))

        if not usuario.check_password(senha):
            flash('Senha incorreta', 'error')
            return redirect(url_for('login'))
            
        login_user(usuario)
        flash('Login realizado com sucesso!', 'success')
        
        if usuario.role == 'Owner':
            return redirect(url_for('owner_dashboard'))
        elif usuario.role == 'Motorista':
            return redirect(url_for('motorista_dashboard'))
        else:
            return redirect(url_for('index'))
            
    return render_template('login.html')

@app.route('/registrar', methods=['GET', 'POST'])
def registrar():
    if request.method == 'POST':
        nome = request.form.get('nome')
        sobrenome = request.form.get('sobrenome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        
        if Usuario.query.filter_by(email=email).first():
            flash('Email já cadastrado', 'error')
            return redirect(url_for('registrar'))
            
        novo_usuario = Usuario(
            nome=nome,
            sobrenome=sobrenome,
            email=email
        )
        novo_usuario.set_password(senha)
        
        try:
            db.session.add(novo_usuario)
            db.session.commit()
            flash('Conta criada com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar conta: {str(e)}', 'error')
            
    return render_template('registrar.html')

@app.route('/promover_admin')
def promover_admin():
    user = Usuario.query.filter_by(email='adminadmin@admin.com').first()
    if user:
        user.is_admin = True
        db.session.commit()
        return "Admin atualizado!"
    return "Usuário não encontrado."


@app.template_filter('dateformat')
def dateformat(value):
    if value:
        return value.strftime('%Y-%m-%d')
    return ''

@app.route('/editar_motorista/<int:motorista_id>', methods=['GET', 'POST'])
def editar_motorista(motorista_id):
    motorista = Motorista.query.get_or_404(motorista_id)
    original_cpf_cnpj = motorista.cpf_cnpj

    if request.method == 'POST':
        if request.form.get('is_modal') == 'true':
            nome = request.form.get('nome', '').strip()
            if not nome:
                flash('O nome é obrigatório.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.nome = nome
            try:
                db.session.commit()
                flash('Nome atualizado com sucesso!', 'success')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            except Exception as e:
                db.session.rollback()
                flash(f'Erro ao atualizar o nome: {str(e)}', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        nome = request.form.get('nome', '').strip()
        data_nascimento = request.form.get('data_nascimento', '').strip()
        endereco = request.form.get('endereco', '').strip()
        pessoa_tipo = request.form.get('pessoa_tipo', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        rg = request.form.get('rg', '').strip() or None
        telefone = request.form.get('telefone', '').strip()
        cnh = request.form.get('cnh', '').strip()
        validade_cnh = request.form.get('validade_cnh', '').strip()
        files = request.files.getlist('anexos')

        if nome:
            if not nome:
                flash('O nome é obrigatório.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.nome = nome

        if data_nascimento:
            try:
                motorista.data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de data de nascimento inválido.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        if endereco:
            if not endereco:
                flash('O endereço é obrigatório.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.endereco = endereco

        if pessoa_tipo:
            if not pessoa_tipo:
                flash('O tipo de pessoa é obrigatório.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.pessoa_tipo = pessoa_tipo

        if cpf_cnpj:
            if not validate_cpf_cnpj(cpf_cnpj, pessoa_tipo or motorista.pessoa_tipo):
                flash(f"{'CPF' if (pessoa_tipo or motorista.pessoa_tipo) == 'fisica' else 'CNPJ'} inválido. Deve conter {'11' if (pessoa_tipo or motorista.pessoa_tipo) == 'fisica' else '14'} dígitos numéricos.", 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            existing_cpf_cnpj = Motorista.query.filter(Motorista.cpf_cnpj == cpf_cnpj, Motorista.id != motorista_id).first()
            if existing_cpf_cnpj:
                flash('CPF/CNPJ já cadastrado.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.cpf_cnpj = cpf_cnpj

        if rg is not None:
            motorista.rg = rg

        if telefone:
            if not validate_telefone(telefone):
                flash('Telefone inválido. Deve conter 10 ou 11 dígitos numéricos.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.telefone = telefone

        if cnh:
            if not validate_cnh(cnh):
                flash('CNH inválida. Deve conter 11 dígitos numéricos.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            existing_cnh = Motorista.query.filter(Motorista.cnh == cnh, Motorista.id != motorista_id).first()
            if existing_cnh:
                flash('CNH já cadastrada.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))
            motorista.cnh = cnh

        if validade_cnh:
            try:
                motorista.validade_cnh = datetime.strptime(validade_cnh, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de validade da CNH inválido.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        if not any([nome, data_nascimento, endereco, pessoa_tipo, cpf_cnpj, rg is not None, telefone, cnh, validade_cnh, files and any(f.filename for f in files)]):
            flash('Nenhum campo foi alterado.', 'error')
            return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        required_fields = ['nome', 'endereco', 'pessoa_tipo', 'cpf_cnpj', 'telefone', 'cnh']
        for field in required_fields:
            if not getattr(motorista, field):
                flash(f'O campo {field} é obrigatório e não pode estar vazio.', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        anexos_urls = motorista.anexos.split(',') if motorista.anexos else []
        allowed_extensions = {'.pdf', '.jpg', '.jpeg', '.png'}
        if files and any(f.filename for f in files):
            try:
                s3_client = boto3.client(
                    's3',
                    endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                    aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                    aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                    region_name='auto'
                )
                bucket_name = app.config['CLOUDFLARE_R2_BUCKET']

                for file in files:
                    if file and file.filename:
                        extension = os.path.splitext(file.filename)[1].lower()
                        if extension not in allowed_extensions:
                            flash(f'Arquivo {file.filename} não permitido. Use PDF, JPG ou PNG.', 'error')
                            continue

                        s3_cpf_cnpj = cpf_cnpj if cpf_cnpj else original_cpf_cnpj
                        filename = secure_filename(file.filename)
                        s3_path = f"motoristas/{s3_cpf_cnpj}/{filename}"

                        s3_client.upload_fileobj(
                            file,
                            bucket_name,
                            s3_path,
                            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'}
                        )
                        public_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
                        anexos_urls.append(public_url)
            except Exception as e:
                flash(f'Erro ao fazer upload dos arquivos: {str(e)}', 'error')
                return redirect(url_for('editar_motorista', motorista_id=motorista_id))

        motorista.anexos = ','.join(anexos_urls) if anexos_urls else None

        try:
            db.session.commit()
            flash('Motorista atualizado com sucesso!', 'success')
            return redirect(url_for('consultar_motoristas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar motorista: {str(e)}', 'error')
            return redirect(url_for('editar_motorista', motorista_id=motorista_id))

    return render_template('editar_motorista.html', motorista=motorista)

@app.route('/owner/dashboard')
@login_required
@owner_required
def owner_dashboard():

    empresas = Empresa.query.options(db.joinedload(Empresa.licenca)).order_by(Empresa.razao_social).all()
    return render_template('owner_dashboard.html', empresas=empresas)

@app.route('/owner/empresa/<int:empresa_id>', methods=['GET', 'POST'])
@login_required
@owner_required
def owner_empresa_detalhes(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)
    # Garante que a empresa tenha uma licença; cria uma se não tiver
    if not empresa.licenca:
        licenca = Licenca(empresa_id=empresa.id)
        db.session.add(licenca)
        db.session.commit()
        # Recarrega a empresa para obter a licença recém-criada
        empresa = Empresa.query.get_or_404(empresa_id)

    if request.method == 'POST':
        try:
            licenca = empresa.licenca
            licenca.plano = request.form.get('plano')
            licenca.max_usuarios = int(request.form.get('max_usuarios'))
            licenca.max_veiculos = int(request.form.get('max_veiculos'))
            data_expiracao_str = request.form.get('data_expiracao')
            licenca.data_expiracao = datetime.strptime(data_expiracao_str, '%Y-%m-%d').date() if data_expiracao_str else None
            licenca.ativo = 'ativo' in request.form

            db.session.commit()
            flash('Licença da empresa atualizada com sucesso!', 'success')
            return redirect(url_for('owner_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar a licença: {e}', 'error')

    return render_template('owner_empresa_detalhes.html', empresa=empresa)


@app.route('/excluir_anexo/<int:motorista_id>/<path:anexo>', methods=['GET'])
def excluir_anexo(motorista_id, anexo):
    motorista = Motorista.query.get_or_404(motorista_id)
    anexos_urls = motorista.anexos.split(',') if motorista.anexos else []
    if anexo in anexos_urls:
        try:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            filename = anexo.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
            try:
                s3_client.delete_object(Bucket=bucket_name, Key=filename)
            except Exception as e:
                logger.error(f"Erro ao excluir anexo {filename}: {str(e)}")
            anexos_urls.remove(anexo)
            motorista.anexos = ','.join(anexos_urls) if anexos_urls else None
            db.session.commit()
            flash('Anexo excluído com sucesso!', 'success')
        except Exception as e:
            flash(f'Erro ao excluir o anexo: {str(e)}', 'error')
    else:
        flash('Anexo não encontrado.', 'error')
    return redirect(url_for('editar_motorista', motorista_id=motorista_id))

@app.route('/excluir_motorista/<int:motorista_id>')
def excluir_motorista(motorista_id):
    motorista = Motorista.query.get_or_404(motorista_id)
    if Viagem.query.filter_by(motorista_id=motorista_id).first():
        flash('Erro: Motorista possui viagens associadas.', 'error')
        return redirect(url_for('consultar_motoristas'))
    try:
        if motorista.anexos:
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            for anexo in motorista.anexos.split(','):
                filename = anexo.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
                try:
                    s3_client.delete_object(Bucket=bucket_name, Key=filename)
                except Exception as e:
                    logger.error(f"Erro ao excluir anexo {filename}: {str(e)}")
        db.session.delete(motorista)
        db.session.commit()
        flash('Motorista excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir motorista: {str(e)}', 'error')
    return redirect(url_for('consultar_motoristas'))

@app.route('/cadastrar_veiculo', methods=['GET', 'POST'])
def cadastrar_veiculo():
    if request.method == 'POST':
        placa = request.form.get('placa', '').strip().upper()
        categoria = request.form.get('categoria', '').strip()
        modelo = request.form.get('modelo', '').strip()
        ano = request.form.get('ano', '').strip()
        valor = request.form.get('valor', '').strip()
        km_rodados = request.form.get('km_rodados', '').strip()
        ultima_manutencao = request.form.get('ultima_manutencao', '').strip()

        if not placa or not modelo:
            flash('Placa e modelo são obrigatórios.', 'error')
            return redirect(url_for('cadastrar_veiculo'))
        if not validate_placa(placa):
            flash('Placa inválida. Deve conter 7 caracteres alfanuméricos.', 'error')
            return redirect(url_for('cadastrar_veiculo'))
        if ano:
            try:
                ano = int(ano)
                if ano < 1900 or ano > datetime.now().year:
                    flash('Ano inválido.', 'error')
                    return redirect(url_for('cadastrar_veiculo'))
            except ValueError:
                flash('Ano deve ser um número válido.', 'error')
                return redirect(url_for('cadastrar_veiculo'))
        if valor:
            try:
                valor = float(valor)
                if valor < 0:
                    flash('Valor deve ser positivo.', 'error')
                    return redirect(url_for('cadastrar_veiculo'))
            except ValueError:
                flash('Valor deve ser um número válido.', 'error')
                return redirect(url_for('cadastrar_veiculo'))
        if km_rodados:
            try:
                km_rodados = float(km_rodados)
                if km_rodados < 0:
                    flash('Km rodados deve ser positivo.', 'error')
                    return redirect(url_for('cadastrar_veiculo'))
            except ValueError:
                flash('Km rodados deve ser um número válido.', 'error')
                return redirect(url_for('cadastrar_veiculo'))
        if ultima_manutencao:
            try:
                ultima_manutencao = datetime.strptime(ultima_manutencao, '%Y-%m-%d').date()
                if ultima_manutencao > datetime.now().date():
                    flash('Data de última manutenção não pode ser no futuro.', 'error')
                    return redirect(url_for('cadastrar_veiculo'))
            except ValueError:
                flash('Formato de data inválido para última manutenção.', 'error')
                return redirect(url_for('cadastrar_veiculo'))

        veiculo = Veiculo(
            placa=placa,
            categoria=categoria or None,
            modelo=modelo,
            ano=ano if ano else None,
            valor=valor if valor else None,
            km_rodados=km_rodados if km_rodados else None,
            ultima_manutencao=ultima_manutencao if ultima_manutencao else None
        )
        try:
            db.session.add(veiculo)
            db.session.commit()
            flash('Veículo cadastrado com sucesso!', 'success')
            return redirect(url_for('index'))
        except IntegrityError:
            db.session.rollback()
            flash('Erro: Placa já cadastrada.', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar veículo: {str(e)}', 'error')
            print(f"Erro ao cadastrar veículo: {str(e)}")
    return render_template('cadastrar_veiculo.html', active_page='cadastrar_veiculo')

@app.route('/consultar_veiculos', methods=['GET'])
def consultar_veiculos():
    search_query = request.args.get('search', '').strip()
    if search_query:
        veiculos = Veiculo.query.filter(
            or_(
                Veiculo.placa.ilike(f'%{search_query}%'),
                Veiculo.modelo.ilike(f'%{search_query}%')
            )
        ).all()
    else:
        veiculos = Veiculo.query.all()
    return render_template('consultar_veiculos.html', veiculos=veiculos, search_query=search_query, active_page='consultar_veiculos')

@app.route('/editar_veiculo/<int:veiculo_id>', methods=['GET', 'POST'])
def editar_veiculo(veiculo_id):
    veiculo = Veiculo.query.get_or_404(veiculo_id)
    if request.method == 'POST':
        placa = request.form.get('placa', '').strip().upper()
        categoria = request.form.get('categoria', '').strip()
        modelo = request.form.get('modelo', '').strip()
        ano = request.form.get('ano', '').strip()
        valor = request.form.get('valor', '').strip()
        km_rodados = request.form.get('km_rodados', '').strip()
        ultima_manutencao = request.form.get('ultima_manutencao', '').strip()

        if not placa or not modelo:
            flash('Placa e modelo são obrigatórios.', 'error')
            return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if not validate_placa(placa):
            flash('Placa inválida. Deve conter 7 caracteres alfanuméricos.', 'error')
            return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if Veiculo.query.filter(Veiculo.placa == placa, Veiculo.id != veiculo_id).first():
            flash('Erro: Placa já cadastrada para outro veículo.', 'error')
            return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if ano:
            try:
                ano = int(ano)
                if ano < 1900 or ano > datetime.now().year:
                    flash('Ano inválido.', 'error')
                    return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
            except ValueError:
                flash('Ano deve ser um número válido.', 'error')
                return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if valor:
            try:
                valor = float(valor)
                if valor < 0:
                    flash('Valor deve ser positivo.', 'error')
                    return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
            except ValueError:
                flash('Valor deve ser um número válido.', 'error')
                return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if km_rodados:
            try:
                km_rodados = float(km_rodados)
                if km_rodados < 0:
                    flash('Km rodados deve ser positivo.', 'error')
                    return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
            except ValueError:
                flash('Km rodados deve ser um número válido.', 'error')
                return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
        if ultima_manutencao:
            try:
                ultima_manutencao = datetime.strptime(ultima_manutencao, '%Y-%m-%d').date()
                if ultima_manutencao > datetime.now().date():
                    flash('Data de última manutenção não pode ser no futuro.', 'error')
                    return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))
            except ValueError:
                flash('Formato de data inválido para última manutenção.', 'error')
                return redirect(url_for('editar_veiculo', veiculo_id=veiculo_id))

        veiculo.placa = placa
        veiculo.categoria = categoria or None
        veiculo.modelo = modelo
        veiculo.ano = ano
        veiculo.valor = valor
        veiculo.km_rodados = km_rodados
        veiculo.ultima_manutencao = ultima_manutencao
        try:
            db.session.commit()
            flash('Veículo atualizado com sucesso!', 'success')
            return redirect(url_for('consultar_veiculos'))
        except:
            db.session.rollback()
            flash('Erro ao atualizar veículo.', 'error')
    return render_template('editar_veiculo.html', veiculo=veiculo, active_page='editar_veiculo')

@app.route('/excluir_veiculo/<int:veiculo_id>')
def excluir_veiculo(veiculo_id):
    veiculo = Veiculo.query.get_or_404(veiculo_id)
    if Viagem.query.filter_by(veiculo_id=veiculo_id).first():
        flash('Erro: Veículo possui viagens associadas.', 'error')
        return redirect(url_for('consultar_veiculos'))
    try:
        db.session.delete(veiculo)
        db.session.commit()
        flash('Veículo excluído com sucesso!', 'success')
    except:
        db.session.rollback()
        flash('Erro ao excluir veículo.', 'error')
    return redirect(url_for('consultar_veiculos'))

@app.route('/iniciar_viagem', methods=['GET', 'POST'])
def iniciar_viagem():
    if request.method == 'POST':
        try:
            motorista_id = request.form.get('motorista_id', '').strip()
            veiculo_id = request.form.get('veiculo_id', '').strip()
            cliente = request.form.get('cliente', '').strip()
            endereco_saida = request.form.get('endereco_saida', '').strip()
            enderecos_destino = request.form.getlist('enderecos_destino[]')
            data_inicio_str = request.form.get('data_inicio', '')
            forma_pagamento = request.form.get('forma_pagamento', '')
            status = request.form.get('status', '')
            observacoes = request.form.get('observacoes', '').strip()

            if not all([motorista_id, veiculo_id, cliente, endereco_saida, enderecos_destino, data_inicio_str, forma_pagamento, status]):
                flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
                return redirect(url_for('iniciar_viagem'))

            if not enderecos_destino:
                flash('Pelo menos um endereço de destino é necessário.', 'error')
                return redirect(url_for('iniciar_viagem'))

            try:
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Formato de data/hora inválido.', 'error')
                return redirect(url_for('iniciar_viagem'))

            veiculo = Veiculo.query.get(veiculo_id)
            if not veiculo:
                flash('Erro: Veículo não encontrado.', 'error')
                return redirect(url_for('iniciar_viagem'))
            if not veiculo.disponivel:
                flash('Erro: Veículo já está em viagem.', 'error')
                return redirect(url_for('iniciar_viagem'))

            enderecos = [endereco_saida] + enderecos_destino
            for endereco in enderecos:
                if not validar_endereco(endereco):
                    flash(f'Endereço inválido: {endereco}. Por favor, insira endereços válidos.', 'error')
                    return redirect(url_for('iniciar_viagem'))

            distancia_km, duracao_segundos = calcular_distancia_e_duracao(enderecos)
            if distancia_km is None or duracao_segundos is None:
                flash('Não foi possível calcular a distância ou duração.', 'error')
                return redirect(url_for('iniciar_viagem'))

            viagem = Viagem(
                motorista_id=motorista_id,
                veiculo_id=veiculo_id,
                cliente=cliente,
                endereco_saida=endereco_saida,
                endereco_destino=enderecos_destino[-1],
                distancia_km=distancia_km,
                data_inicio=data_inicio,
                duracao_segundos=duracao_segundos,
                forma_pagamento=forma_pagamento,
                status=status,
                observacoes=observacoes or None
            )
            veiculo.disponivel = False
            db.session.add(viagem)
            db.session.flush()

            for ordem, endereco in enumerate(enderecos_destino, 1):
                destino = Destino(
                    viagem_id=viagem.id,
                    endereco=endereco,
                    ordem=ordem
                )
                db.session.add(destino)

            db.session.commit()
            flash(f'Viagem iniciada com sucesso! Distância: {distancia_km:.2f} km, Duração estimada: {duracao_segundos//60} minutos', 'success')
            return redirect(url_for('iniciar_viagem'))
        except Exception as e:
            logger.error(f"Erro ao iniciar viagem: {str(e)}")
            db.session.rollback()
            flash(f'Erro ao iniciar viagem: {str(e)}', 'error')
            return redirect(url_for('iniciar_viagem'))

    motoristas = Motorista.query.all()
    veiculos = Veiculo.query.filter_by(disponivel=True).all()
    viagens = Viagem.query.filter_by(data_fim=None).order_by(Viagem.data_inicio.desc()).all()
    viagens_data = []
    for viagem in viagens:
        data_inicio_formatada = viagem.data_inicio.strftime('%Y-%m-%d %H:%M:%S')[:-3]
        horario_chegada = (viagem.data_inicio + timedelta(seconds=viagem.duracao_segundos)).strftime('%d/%m/%Y %H:%M') if viagem.duracao_segundos else 'Não calculado'
        
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        viagem_dict = {
            'valor_recebido': viagem.valor_recebido or 0,
            'id': viagem.id,
            'motorista_id': viagem.motorista_id,
            'motorista_cpf_cnpj': viagem.motorista_cpf_cnpj,
            'veiculo_id': viagem.veiculo_id,
            'cliente': viagem.cliente,
            'endereco_saida': viagem.endereco_saida,
            'endereco_destino': viagem.endereco_destino,
            'data_inicio': data_inicio_formatada,
            'duracao_segundos': viagem.duracao_segundos,
            'custo': viagem.custo,
            'forma_pagamento': viagem.forma_pagamento,
            'status': viagem.status,
            'observacoes': viagem.observacoes,
            'motorista_nome': motorista_nome,
            'veiculo_placa': viagem.veiculo.placa,
            'veiculo_modelo': viagem.veiculo.modelo,
            'distancia_km': viagem.distancia_km,
            'horario_chegada': horario_chegada,
            'destinos': [{'endereco': destino.endereco, 'ordem': destino.ordem} for destino in viagem.destinos]
        }
        viagens_data.append(viagem_dict)
    return render_template('iniciar_viagem.html', motoristas=motoristas, veiculos=veiculos, viagens=viagens_data, Maps_API_KEY=Maps_API_KEY)

@app.route('/editar_viagem/<int:viagem_id>', methods=['GET', 'POST'])
def editar_viagem(viagem_id):
    viagem = Viagem.query.get_or_404(viagem_id)
    if request.method == 'POST':
        try:
            motorista_id = request.form.get('motorista_id', '').strip()
            veiculo_id = request.form.get('veiculo_id', '').strip()
            cliente = request.form.get('cliente', '').strip()
            endereco_saida = request.form.get('endereco_saida', '').strip()
            enderecos_destino = request.form.getlist('enderecos_destino[]')
            data_inicio_str = request.form.get('data_inicio', '')
            forma_pagamento = request.form.get('forma_pagamento', '')
            status = request.form.get('status', '')
            observacoes = request.form.get('observacoes', '').strip()

            if not all([veiculo_id, cliente, endereco_saida, enderecos_destino, data_inicio_str, forma_pagamento, status]):
                flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
                return redirect(url_for('iniciar_viagem'))

            if not enderecos_destino:
                flash('Pelo menos um endereço de destino é necessário.', 'error')
                return redirect(url_for('iniciar_viagem'))

            try:
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('Formato de data/hora inválido.', 'error')
                return redirect(url_for('iniciar_viagem'))

            enderecos = [endereco_saida] + enderecos_destino
            for endereco in enderecos:
                if not validar_endereco(endereco):
                    flash(f'Endereço inválido: {endereco}. Por favor, insira endereços válidos.', 'error')
                    return redirect(url_for('iniciar_viagem'))

            distancia_km, duracao_segundos = calcular_distancia_e_duracao(enderecos)
            if distancia_km is None or duracao_segundos is None:
                flash('Não foi possível recalcular a distância ou duração.', 'error')
                return redirect(url_for('iniciar_viagem'))

            viagem.motorista_id = motorista_id if motorista_id else None # Definir como None se vazio
            viagem.veiculo_id = veiculo_id
            viagem.cliente = cliente
            viagem.endereco_saida = endereco_saida
            viagem.endereco_destino = enderecos_destino[-1]
            viagem.data_inicio = data_inicio
            viagem.distancia_km = distancia_km
            viagem.duracao_segundos = duracao_segundos
            viagem.forma_pagamento = forma_pagamento
            viagem.status = status
            viagem.observacoes = observacoes or None

            Destino.query.filter_by(viagem_id=viagem.id).delete()
            for ordem, endereco in enumerate(enderecos_destino, 1):
                destino = Destino(
                    viagem_id=viagem.id,
                    endereco=endereco,
                    ordem=ordem
                )
                db.session.add(destino)

            db.session.commit()
            flash('Viagem atualizada com sucesso!', 'success')
            return redirect(url_for('iniciar_viagem'))
        except Exception as e:
            logger.error(f"Erro ao editar viagem: {str(e)}")
            db.session.rollback()
            flash(f'Erro ao editar viagem: {str(e)}', 'error')
            return redirect(url_for('iniciar_viagem'))

    motoristas = Motorista.query.all()
    veiculos = Veiculo.query.filter_by(disponivel=True).all()
    if viagem.veiculo.disponivel == False:
        veiculos.append(viagem.veiculo)
    viagens = Viagem.query.filter_by(data_fim=None).order_by(Viagem.data_inicio.desc()).all()
    viagens_data = []
    for v in viagens:
        horario_chegada = (v.data_inicio + timedelta(seconds=v.duracao_segundos)).strftime('%d/%m/%Y %H:%M') if v.duracao_segundos else 'Não calculado'
        
        motorista_nome = 'N/A'
        if v.motorista_id:
            motorista_nome = v.motorista_formal.nome if v.motorista_formal else 'N/A'
        elif v.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=v.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=v.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        viagem_dict = {
            'id': v.id,
            'motorista_id': v.motorista_id,
            'motorista_cpf_cnpj': v.motorista_cpf_cnpj,
            'veiculo_id': v.veiculo_id,
            'cliente': v.cliente,
            'endereco_saida': v.endereco_saida,
            'endereco_destino': v.endereco_destino,
            'data_inicio': v.data_inicio.strftime('%Y-%m-%dT%H:%M:%S'),
            'duracao_segundos': v.duracao_segundos,
            'custo': v.custo,
            'forma_pagamento': v.forma_pagamento,
            'status': v.status,
            'observacoes': v.observacoes,
            'motorista_nome': motorista_nome,
            'veiculo_placa': v.veiculo.placa,
            'veiculo_modelo': v.veiculo.modelo,
            'distancia_km': v.distancia_km,
            'horario_chegada': horario_chegada,
            'destinos': [{'endereco': destino.endereco, 'ordem': destino.ordem} for destino in v.destinos]
        }
        viagens_data.append(viagem_dict)
    return render_template('iniciar_viagem.html', motoristas=motoristas, veiculos=veiculos, viagens=viagens_data, viagem_edit=viagem, Maps_API_KEY=Maps_API_KEY)

@app.route('/excluir_viagem/<int:viagem_id>')
def excluir_viagem(viagem_id):
    viagem = Viagem.query.get_or_404(viagem_id)
    try:
        if not viagem.data_fim:
            viagem.veiculo.disponivel = True
        db.session.delete(viagem)
        db.session.commit()
        flash('Viagem excluída com sucesso!', 'success')
    except Exception as e:
        logger.error(f"Erro ao excluir viagem: {str(e)}")
        flash(f'Erro ao excluir viagem: {str(e)}', 'error')
    return redirect(url_for('index'))

@app.route('/salvar_custo_viagem', methods=['POST'])
@login_required
def salvar_custo_viagem():
    viagem_id = request.form.get('viagem_id')
    if not viagem_id:
        return jsonify({'success': False, 'message': 'ID da viagem não fornecido.'}), 400

    try:
        viagem = Viagem.query.get_or_404(viagem_id)
        custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        if not custo:
            custo = CustoViagem(viagem_id=viagem_id)
            db.session.add(custo)

        custo.pedagios = float(request.form.get('pedagios') or 0)
        custo.alimentacao = float(request.form.get('alimentacao') or 0)
        custo.hospedagem = float(request.form.get('hospedagem') or 0)
        custo.outros = float(request.form.get('outros') or 0)
        custo.descricao_outros = request.form.get('descricao_outros')

        files = request.files.getlist('anexos_despesa')
        anexos_urls = custo.anexos.split(',') if custo.anexos else []
        
        if files and any(f.filename for f in files):
            s3_client = boto3.client(
                's3',
                endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
                aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
                aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
                region_name='auto'
            )
            bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
            for file in files:
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    s3_path = f"custos_viagem/{viagem_id}/{uuid.uuid4()}-{filename}"
                    s3_client.upload_fileobj(
                        file, bucket_name, s3_path,
                        ExtraArgs={'ContentType': file.content_type}
                    )
                    public_url = f"{app.config['CLOUDFLARE_R2_PUBLIC_URL']}/{s3_path}"
                    anexos_urls.append(public_url)

        custo.anexos = ','.join(anexos_urls) if anexos_urls else None

        custo_total = (custo.pedagios or 0) + (custo.alimentacao or 0) + (custo.hospedagem or 0) + (custo.outros or 0)
        viagem.custo = custo_total
        
        db.session.commit()

        return jsonify({
            'success': True, 
            'message': 'Despesas salvas com sucesso!',
            'custo_total_formatado': f'R$ {custo_total:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            'anexos': anexos_urls
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar custo da viagem {viagem_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro interno do servidor: {str(e)}'}), 500
    

@app.route('/excluir_anexo_custo', methods=['POST'])
@login_required
def excluir_anexo_custo():
    data = request.get_json()
    viagem_id = data.get('viagem_id')
    anexo_url = data.get('anexo_url')

    if not viagem_id or not anexo_url:
        return jsonify({'success': False, 'message': 'Dados incompletos.'}), 400

    try:
        custo = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        if not custo or not custo.anexos:
            return jsonify({'success': False, 'message': 'Anexo não encontrado.'}), 404

        anexos_atuais = custo.anexos.split(',')
        if anexo_url not in anexos_atuais:
            return jsonify({'success': False, 'message': 'URL do anexo não corresponde.'}), 404

        # 1. Excluir do Cloudflare R2
        s3_client = boto3.client(
            's3',
            endpoint_url=app.config['CLOUDFLARE_R2_ENDPOINT'],
            aws_access_key_id=app.config['CLOUDFLARE_R2_ACCESS_KEY'],
            aws_secret_access_key=app.config['CLOUDFLARE_R2_SECRET_KEY'],
            region_name='auto'
        )
        bucket_name = app.config['CLOUDFLARE_R2_BUCKET']
        key = anexo_url.replace(app.config['CLOUDFLARE_R2_PUBLIC_URL'] + '/', '')
        s3_client.delete_object(Bucket=bucket_name, Key=key)

        # 2. Excluir do Banco de Dados
        anexos_atuais.remove(anexo_url)
        custo.anexos = ','.join(anexos_atuais) if anexos_atuais else None
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Anexo excluído com sucesso!'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao excluir anexo da viagem {viagem_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro interno do servidor: {str(e)}'}), 500

@app.route('/consultar_despesas/<int:viagem_id>', methods=['GET'])
def consultar_despesas(viagem_id):
    try:
        viagem = Viagem.query.get_or_404(viagem_id)
        custo_viagem = CustoViagem.query.filter_by(viagem_id=viagem_id).first()
        custo_dict = {
            'combustivel': custo_viagem.combustivel if custo_viagem else 0.0,
            'pedagios': custo_viagem.pedagios if custo_viagem else 0.0,
            'alimentacao': custo_viagem.alimentacao if custo_viagem else 0.0,
            'hospedagem': custo_viagem.hospedagem if custo_viagem else 0.0,
            'outros': custo_viagem.outros if custo_viagem else 0.0,
            'descricao_outros': custo_viagem.descricao_outros if custo_viagem else 'Nenhuma',
            'anexos': custo_viagem.anexos.split(',') if custo_viagem and custo_viagem.anexos else []
        }
        return jsonify(custo_dict)
    except Exception as e:
        logger.error(f"Erro ao consultar despesas da viagem {viagem_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/atualizar_status_viagem/<int:viagem_id>', methods=['POST'])
def atualizar_status_viagem(viagem_id):
    try:
        data = request.get_json()
        novo_status = data.get('status')

        if novo_status not in ['pendente', 'em_andamento', 'concluida', 'cancelada']:
            return jsonify({'success': False, 'message': 'Status inválido'}), 400

        viagem = Viagem.query.get_or_404(viagem_id)
        viagem.status = novo_status

        if novo_status == 'concluida' and not viagem.data_fim:
            viagem.data_fim = datetime.now()

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao atualizar status da viagem {viagem_id}: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# Em app.py, SUBSTITUA a função consultar_viagens inteira por esta:

@app.route('/consultar_viagens')
@login_required # Adicionado para proteger a rota
def consultar_viagens():
    # 1. Captura de todos os filtros da URL
    status_filter = request.args.get('status', '')
    search_query = request.args.get('search', '')
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')
    motorista_id_filter = request.args.get('motorista_id', type=int)
    veiculo_id_filter = request.args.get('veiculo_id', type=int)

    query = Viagem.query

    # 2. Aplicação dos filtros na query do SQLAlchemy
    if status_filter:
        query = query.filter(Viagem.status == status_filter)
    if motorista_id_filter:
        query = query.filter(Viagem.motorista_id == motorista_id_filter)
    if veiculo_id_filter:
        query = query.filter(Viagem.veiculo_id == veiculo_id_filter)

    if search_query:
        query = query.outerjoin(Motorista, Viagem.motorista_id == Motorista.id).outerjoin(Usuario, Motorista.usuario_id == Usuario.id).filter(
            (Viagem.cliente.ilike(f'%{search_query}%')) |
            (Usuario.nome.ilike(f'%{search_query}%')) |
            (Motorista.nome.ilike(f'%{search_query}%')) |
            (Viagem.veiculo.has(Veiculo.placa.ilike(f'%{search_query}%'))) |
            (Viagem.endereco_saida.ilike(f'%{search_query}%')) |
            (Viagem.endereco_destino.ilike(f'%{search_query}%'))
        )
        
    if data_inicio:
        try:
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        except ValueError:
            flash('Data de início inválida.', 'error')
            
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1, seconds=-1)
            query = query.filter(Viagem.data_inicio <= data_fim_obj)
        except ValueError:
            flash('Data de fim inválida.', 'error')

    # Executa a query final
    viagens_objetos = query.order_by(Viagem.data_inicio.desc()).all()

    # Adiciona o nome do motorista a cada objeto, para facilitar no template
    for v in viagens_objetos:
        v.motorista_nome = 'N/A' # Atributo temporário
        if v.motorista_id:
            v.motorista_nome = v.motorista_formal.nome if v.motorista_formal else 'N/A'
        elif v.motorista_cpf_cnpj:
            usuario = Usuario.query.filter_by(cpf_cnpj=v.motorista_cpf_cnpj).first()
            if usuario:
                v.motorista_nome = f"{usuario.nome} {usuario.sobrenome}"

    # 3. Busca de dados para os menus de filtro
    motoristas_filtro = Motorista.query.order_by(Motorista.nome).all()
    veiculos_filtro = Veiculo.query.order_by(Veiculo.placa).all()
    
    # 4. Renderização do template com os dados necessários
    return render_template(
        'consultar_viagens.html',
        active_page='consultar_viagens',
        viagens=viagens_objetos,  # Passando a lista de OBJETOS Viagem
        motoristas=motoristas_filtro,
        veiculos=veiculos_filtro,
        # Passando os valores dos filtros de volta para manter os campos preenchidos
        request=request # Passa o objeto request inteiro para fácil acesso no template
    )

from flask import jsonify, request
from datetime import datetime

@app.route('/finalizar_viagem/<int:viagem_id>', methods=['POST'])
@login_required
def finalizar_viagem(viagem_id):
    viagem = Viagem.query.get_or_404(viagem_id)

    try:
        viagem.data_fim = datetime.utcnow()
        viagem.status = 'concluida'
        viagem.veiculo.disponivel = True

        # Aqui você pode capturar o valor_recebido da requisição se quiser
        data = request.get_json()
        if data and 'valor_recebido' in data:
            viagem.valor_recebido = float(data['valor_recebido'])

        db.session.commit()

        return jsonify(success=True, message='Viagem finalizada com sucesso.')
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500

    

from collections import defaultdict

from collections import defaultdict

# Em app.py, substitua a função relatorios inteira por esta:

@app.route('/relatorios')
@login_required
def relatorios():
    try:
        data_inicio_str = request.args.get('data_inicio', '')
        data_fim_str = request.args.get('data_fim', '')
        motorista_id_filter = request.args.get('motorista_id', '')
        veiculo_id_filter = request.args.get('veiculo_id', '')
        
        query = Viagem.query.options(
            db.joinedload(Viagem.custo_viagem),
            db.joinedload(Viagem.motorista_formal),
            db.joinedload(Viagem.veiculo),
            db.joinedload(Viagem.abastecimentos) # Carrega os abastecimentos junto
        )

        if data_inicio_str:
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio_str, '%Y-%m-%d'))
        if data_fim_str:
            data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Viagem.data_inicio < data_fim_obj)
        if motorista_id_filter:
            query = query.filter(Viagem.motorista_id == int(motorista_id_filter))
        if veiculo_id_filter:
            query = query.filter(Viagem.veiculo_id == int(veiculo_id_filter))

        viagens_filtradas = query.order_by(Viagem.data_inicio.desc()).all()

        total_receita = 0.0
        total_custo_outros = 0.0
        total_custo_combustivel = 0.0
        total_distancia = 0.0
        total_litros = 0.0
        
        dados_grafico_mensal = defaultdict(lambda: {'receita': 0.0, 'custo': 0.0})
        dados_grafico_categorias = defaultdict(float)
        clientes_stats = defaultdict(lambda: {'viagens': 0, 'receita': 0.0})
        motoristas_stats = defaultdict(lambda: {'id': None, 'nome': 'N/A', 'viagens': 0, 'receita': 0.0})
        # Alterado para coletar mais dados
        veiculos_stats = defaultdict(lambda: {'id': None, 'modelo': 'N/A', 'placa': 'N/A', 'km': 0.0, 'custo': 0.0, 'litros': 0.0})

        for v in viagens_filtradas:
            receita_viagem = v.valor_recebido or 0.0

            # Calcula custo de combustível REAL a partir dos abastecimentos
            custo_combustivel_viagem = sum(a.custo_total for a in v.abastecimentos)
            litros_viagem = sum(a.litros for a in v.abastecimentos)

            # Calcula outros custos (sem combustível)
            custo_outros_viagem = 0
            if v.custo_viagem:
                custo_outros_viagem += (v.custo_viagem.pedagios or 0)
                custo_outros_viagem += (v.custo_viagem.alimentacao or 0)
                custo_outros_viagem += (v.custo_viagem.hospedagem or 0)
                custo_outros_viagem += (v.custo_viagem.outros or 0)
            
            custo_total_viagem = custo_combustivel_viagem + custo_outros_viagem

            total_receita += receita_viagem
            total_custo_combustivel += custo_combustivel_viagem
            total_custo_outros += custo_outros_viagem
            total_distancia += v.distancia_km or 0.0
            total_litros += litros_viagem

            if v.data_inicio:
                mes = v.data_inicio.strftime('%Y-%m')
                dados_grafico_mensal[mes]['receita'] += receita_viagem
                dados_grafico_mensal[mes]['custo'] += custo_total_viagem

            # Preenche dados para o gráfico de categorias
            dados_grafico_categorias['Combustível'] += custo_combustivel_viagem
            if v.custo_viagem:
                dados_grafico_categorias['Pedágios'] += v.custo_viagem.pedagios or 0
                dados_grafico_categorias['Alimentação'] += v.custo_viagem.alimentacao or 0
                dados_grafico_categorias['Hospedagem'] += v.custo_viagem.hospedagem or 0
                dados_grafico_categorias['Outros'] += v.custo_viagem.outros or 0
            
            if v.cliente:
                clientes_stats[v.cliente]['viagens'] += 1
                clientes_stats[v.cliente]['receita'] += receita_viagem

            if v.motorista_formal:
                motoristas_stats[v.motorista_formal.id].update({
                    'id': v.motorista_formal.id, 'nome': v.motorista_formal.nome
                })
                motoristas_stats[v.motorista_formal.id]['viagens'] += 1
                motoristas_stats[v.motorista_formal.id]['receita'] += receita_viagem

            if v.veiculo:
                veiculos_stats[v.veiculo.id].update({
                    'id': v.veiculo.id, 'placa': v.veiculo.placa, 'modelo': v.veiculo.modelo
                })
                veiculos_stats[v.veiculo.id]['km'] += v.distancia_km or 0.0
                veiculos_stats[v.veiculo.id]['custo'] += custo_total_viagem
                veiculos_stats[v.veiculo.id]['litros'] += litros_viagem
        
        # Consolida o custo total e calcula a média geral de consumo
        total_custo = total_custo_combustivel + total_custo_outros
        consumo_medio_geral = (total_distancia / total_litros) if total_litros > 0 else 0

        motoristas_para_filtro = Motorista.query.order_by(Motorista.nome).all()
        veiculos_para_filtro = Veiculo.query.order_by(Veiculo.placa).all()

        return render_template(
            'relatorios.html',
            request=request,
            total_viagens=len(viagens_filtradas),
            total_receita=total_receita,
            total_custo=total_custo,
            consumo_medio_geral=consumo_medio_geral,  # <-- NOVO DADO
            motoristas_filtro=motoristas_para_filtro,
            veiculos_filtro=veiculos_para_filtro,
            dados_grafico_mensal=dict(sorted(dados_grafico_mensal.items())),
            dados_grafico_categorias=dict(dados_grafico_categorias),
            clientes_stats=list(clientes_stats.values()),
            motoristas=motoristas_stats,
            veiculos_stats=list(veiculos_stats.values()) # <-- NOVO DADO
        )

    except Exception as e:
        logger.error(f"Erro ao gerar relatórios: {e}", exc_info=True)
        flash(f"Ocorreu um erro inesperado ao gerar os relatórios: {e}", "error")
        return redirect(url_for('index'))
    
@app.route('/exportar_relatorio')
def exportar_relatorio():
    try:
        data_inicio = request.args.get('data_inicio', '')
        data_fim = request.args.get('data_fim', '')
        motorista_id_filter = request.args.get('motorista_id', '') # Renomeado
        status_filter = request.args.get('status', '')

        query = Viagem.query

        if data_inicio:
            query = query.filter(Viagem.data_inicio >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            query = query.filter(Viagem.data_inicio <= datetime.strptime(data_fim, '%Y-%m-%d'))
        if motorista_id_filter:
            query = query.filter_by(motorista_id=motorista_id_filter)
        if status_filter:
            query = query.filter_by(status=status_filter)

        viagens = query.outerjoin(Motorista).outerjoin(Veiculo).all() # Usar outerjoin para não excluir viagens sem motorista formal

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatório Financeiro"

        headers = [
            "ID", "Data", "Cliente", "Motorista", "Veículo",
            "Distância (km)", "Receita (R$)", "Custo (R$)", "Lucro (R$)",
            "Forma Pagamento", "Status"
        ]
        
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header).font = Font(bold=True)

        for row_num, viagem in enumerate(viagens, 2):
            receita = viagem.valor_recebido or 0
            custo = viagem.custo or 0
            lucro = receita - custo
            
            motorista_nome = 'N/A'
            if viagem.motorista_id:
                motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
            elif viagem.motorista_cpf_cnpj:
                usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if usuario_com_cpf:
                    motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
                else:
                    motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                    if motorista_formal_cpf:
                        motorista_nome = motorista_formal_cpf.nome
            
            veiculo_info = f"{viagem.veiculo.placa} - {viagem.veiculo.modelo}" if viagem.veiculo else 'N/A'

            sheet.cell(row=row_num, column=1, value=viagem.id)
            sheet.cell(row=row_num, column=2, value=viagem.data_inicio.strftime('%d/%m/%Y'))
            sheet.cell(row=row_num, column=3, value=viagem.cliente)
            sheet.cell(row=row_num, column=4, value=motorista_nome) # Usar o nome processado
            sheet.cell(row=row_num, column=5, value=veiculo_info) # Usar info do veículo processada
            sheet.cell(row=row_num, column=6, value=viagem.distancia_km or 0)
            sheet.cell(row=row_num, column=7, value=receita)
            sheet.cell(row=row_num, column=8, value=custo)
            sheet.cell(row=row_num, column=9, value=lucro)
            sheet.cell(row=row_num, column=10, value=viagem.forma_pagamento or '')
            sheet.cell(row=row_num, column=11, value=viagem.status)

        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"relatorio_financeiro_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Erro ao exportar relatório: {str(e)}", exc_info=True)
        flash('Erro ao gerar relatório em Excel', 'error')
        return redirect(url_for('relatorios'))

@app.route('/get_active_trip')
def get_active_trip():
    viagem = Viagem.query.filter_by(data_fim=None, status='em_andamento').first()
    if viagem:
        horario_chegada = (viagem.data_inicio + timedelta(seconds=viagem.duracao_segundos)).strftime('%d/%m/%Y %H:%M') if viagem.duracao_segundos else 'Não calculado'
        
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        trip_data = {
            'trip': {
                'motorista_nome': motorista_nome,
                'veiculo_placa': viagem.veiculo.placa if viagem.veiculo else 'N/A',
                'veiculo_modelo': viagem.veiculo.modelo if viagem.veiculo else 'N/A',
                'endereco_saida': viagem.endereco_saida,
                'endereco_destino': viagem.endereco_destino,
                'horario_chegada': horario_chegada
            }
        }
        return jsonify(trip_data)
    return jsonify({'trip': None})

@app.route('/create_admin')
def create_admin():
    if not Usuario.query.filter_by(email='adminadmin@admin.com').first():
        admin = Usuario(
            nome='Admin',
            sobrenome='Admin',
            email='adminadmin@admin.com',
            telefone='11999999999',
            role='Admin',
            is_admin=True,
            cpf_cnpj='00000000000' # CPF/CNPJ padrão para admin
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        return 'Usuário admin criado!'
    return 'Usuário já existe'


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu com sucesso.', 'success')
    return redirect(url_for('login'))

@app.route('/configuracoes', methods=['GET', 'POST'])
@login_required
def configuracoes():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        idioma = request.form.get('idioma', '').strip()

        if not nome or not sobrenome:
            flash('Nome e sobrenome são obrigatórios.', 'error')
            return redirect(url_for('configuracoes'))
        
        if idioma not in ['Português', 'Inglês', 'Espanhol']:
            flash('Idioma inválido.', 'error')
            return redirect(url_for('configuracoes'))

        current_user.nome = nome
        current_user.sobrenome = sobrenome
        current_user.idioma = idioma

        try:
            db.session.commit()
            flash('Configurações pessoais atualizadas com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar configurações: {str(e)}', 'error')

        return redirect(url_for('configuracoes'))

    usuarios = []
    empresa = None

    if current_user.empresa_id:
        empresa = db.session.get(Empresa, current_user.empresa_id)

    if current_user.is_admin and current_user.empresa_id:
        usuarios = Usuario.query.filter_by(empresa_id=current_user.empresa_id).all()
    elif current_user.is_admin:
        usuarios = [current_user]

    return render_template('configuracoes.html', usuarios=usuarios, empresa=empresa)


@app.route('/editar_usuario/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        sobrenome = request.form.get('sobrenome', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', '').strip()
        senha = request.form.get('senha', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip() # Pega CPF/CNPJ do form

        if not nome or not sobrenome or not email or not role:
            flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        if role not in ['Motorista', 'Master', 'Admin']:
            flash('Papel inválido.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        if email != usuario.email and Usuario.query.filter_by(email=email).first():
            flash('E-mail já cadastrado.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))
        
        if cpf_cnpj and cpf_cnpj != usuario.cpf_cnpj and Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first():
            flash('CPF/CNPJ já cadastrado para outro usuário.', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

        usuario.nome = nome
        usuario.sobrenome = sobrenome
        usuario.email = email
        usuario.role = role
        usuario.is_admin = (role == 'Admin')
        usuario.cpf_cnpj = cpf_cnpj if cpf_cnpj else None # Atualiza CPF/CNPJ do usuário

        if senha:
            usuario.set_password(senha)

        try:
            db.session.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('configuracoes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar usuário: {str(e)}', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))

    return render_template('editar_usuario.html', usuario=usuario)




@app.route('/gerenciar_empresa', methods=['GET', 'POST'])
@login_required
def gerenciar_empresa():
    
    if current_user.role not in ['Admin', 'Master']:
        flash('Acesso negado. Apenas administradores podem gerenciar a empresa.', 'error')
        return redirect(url_for('index'))

    empresa = db.session.get(Empresa, current_user.empresa_id) if current_user.empresa_id else None

    if request.method == 'POST':
        cnpj = re.sub(r'\D', '', request.form.get('cnpj', '')) 

        if not validate_cpf_cnpj(cnpj, 'juridica'):
            flash('CNPJ inválido. Deve conter 14 dígitos numéricos.', 'error')
            
            return render_template('gerenciar_empresa.html', empresa=request.form)

        
        empresa_existente = Empresa.query.filter(Empresa.cnpj == cnpj).first()
        if empresa_existente and (not empresa or empresa.id != empresa_existente.id):
            flash('Este CNPJ já está cadastrado em outra empresa.', 'error')
            return render_template('gerenciar_empresa.html', empresa=request.form)

        if empresa:
            # --- LÓGICA DE ATUALIZAÇÃO ---
            empresa.razao_social = request.form.get('razao_social').strip()
            empresa.nome_fantasia = request.form.get('nome_fantasia').strip()
            empresa.cnpj = cnpj
            empresa.inscricao_estadual = request.form.get('inscricao_estadual').strip()
            empresa.endereco = request.form.get('endereco').strip()
            empresa.cidade = request.form.get('cidade').strip()
            empresa.estado = request.form.get('estado').strip().upper()
            empresa.cep = re.sub(r'\D', '', request.form.get('cep', ''))
            empresa.telefone = re.sub(r'\D', '', request.form.get('telefone', ''))
            empresa.email_contato = request.form.get('email_contato').strip()
            flash('Dados da empresa atualizados com sucesso!', 'success')
        else:
            # --- LÓGICA DE CRIAÇÃO ---
            nova_empresa = Empresa(
                razao_social=request.form.get('razao_social').strip(),
                nome_fantasia=request.form.get('nome_fantasia').strip(),
                cnpj=cnpj,
                inscricao_estadual=request.form.get('inscricao_estadual').strip(),
                endereco=request.form.get('endereco').strip(),
                cidade=request.form.get('cidade').strip(),
                estado=request.form.get('estado').strip().upper(),
                cep=re.sub(r'\D', '', request.form.get('cep', '')),
                telefone=re.sub(r'\D', '', request.form.get('telefone', '')),
                email_contato=request.form.get('email_contato').strip()
            )
            db.session.add(nova_empresa)
            db.session.flush() 

            
            current_user.empresa_id = nova_empresa.id
            flash('Empresa cadastrada com sucesso!', 'success')
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao salvar os dados: {e}', 'error')
            return render_template('gerenciar_empresa.html', empresa=request.form)
            
        return redirect(url_for('configuracoes'))

    # --- LÓGICA GET ---
    # Mostra o formulário preenchido para edição ou vazio para criação
    return render_template('gerenciar_empresa.html', empresa=empresa)


@app.route('/excluir_usuario/<int:usuario_id>')
@login_required
@admin_required
def excluir_usuario(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    if usuario.id == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'error')
        return redirect(url_for('configuracoes'))

    try:
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuário excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir usuário: {str(e)}', 'error')
    return redirect(url_for('configuracoes'))

def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or (current_user.role not in ['Admin', 'Master']):
            flash('Acesso restrito a administradores ou masters.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/motorista_dashboard')
@login_required
def motorista_dashboard():
    if hasattr(current_user, 'motorista') and current_user.motorista:
        if isinstance(current_user.motorista, list):
            motorista = current_user.motorista[0] if current_user.motorista else None
        else:
            motorista = current_user.motorista
        motorista_id = motorista.id if motorista else None
    else:
        # Busca o motorista formal vinculado pelo cpf_cnpj, se existir
        motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
        motorista_id = motorista_formal.id if motorista_formal else None

    viagens = Viagem.query.filter(
        or_(
            Viagem.motorista_cpf_cnpj == current_user.cpf_cnpj,
            Viagem.motorista_id == motorista_id
        )
    ).all()

    viagem_ativa = next((v for v in viagens if v.status == 'em_andamento'), None)

    viagens_data = []
    for viagem in viagens:
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario_com_cpf:
                motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
            else:
                motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
                if motorista_formal_cpf:
                    motorista_nome = motorista_formal_cpf.nome

        viagens_data.append({
            'id': viagem.id,
            'cliente': viagem.cliente,
            'motorista_nome': motorista_nome,
            'endereco_saida': viagem.endereco_saida,
            'endereco_destino': viagem.endereco_destino,
            'status': viagem.status,
            'veiculo_placa': viagem.veiculo.placa,
            'veiculo_modelo': viagem.veiculo.modelo,
            'data_inicio': viagem.data_inicio,
            'data_fim': viagem.data_fim,
            'distancia_km': viagem.distancia_km
        })

    return render_template(
        'motorista_dashboard.html',
        viagens=viagens_data,
        viagem_ativa=viagem_ativa,
    )





@app.route('/atualizar_localizacao', methods=['POST'])
@login_required
def atualizar_localizacao():
    data = request.get_json()
    lat = data.get('latitude')
    lon = data.get('longitude')
    viagem_id = data.get('viagem_id')

    if not lat or not lon:
        return jsonify({'success': False, 'message': 'Coordenadas inválidas.'})

    try:
        endereco = get_address_geoapify(lat, lon)

        # Buscar o motorista formal vinculado ao usuário logado pelo cpf_cnpj
        motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
        motorista_id_para_localizacao = motorista_formal.id if motorista_formal else None

        if not motorista_id_para_localizacao:
            logger.warning(f"Usuário {current_user.email} tentou atualizar localização sem motorista formal vinculado por CPF/CNPJ.")
            return jsonify({'success': False, 'message': 'Motorista formal não encontrado para vincular localização.'})


        nova_localizacao = Localizacao(
            motorista_id=motorista_id_para_localizacao,
            viagem_id=viagem_id,
            latitude=lat,
            longitude=lon,
            endereco=endereco
        )
        db.session.add(nova_localizacao)
        db.session.commit()

        return jsonify({'success': True, 'endereco': endereco})
    except Exception as e:
        logger.error(f"Erro ao atualizar localização: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)})


@app.route('/selecionar_viagem/<int:viagem_id>', methods=['POST'])
@login_required
def selecionar_viagem(viagem_id):
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso negado'})
    
    if not current_user.cpf_cnpj:
        return jsonify({'success': False, 'message': 'Seu perfil de usuário não possui CPF/CNPJ. Preencha-o nas configurações para iniciar viagens.'})

    viagem = Viagem.query.get(viagem_id)

    if not viagem:
        return jsonify({'success': False, 'message': 'Viagem não encontrada'})

    if viagem.status != 'pendente': # 'Pendente' precisa ser 'pendente' conforme o default do modelo
        return jsonify({'success': False, 'message': 'Viagem já foi iniciada ou está em outro status'})

    viagem.motorista_cpf_cnpj = current_user.cpf_cnpj # Vincula pelo CPF/CNPJ do usuário
    
    # Opcional: Se o usuário logado tiver um motorista formal vinculado, use o ID desse motorista também
    motorista_formal = Motorista.query.filter_by(usuario_id=current_user.id, cpf_cnpj=current_user.cpf_cnpj).first()
    if motorista_formal:
        viagem.motorista_id = motorista_formal.id # Linka com o ID do motorista formal se ele existir

    viagem.status = 'em_andamento' # 'Ativa' precisa ser 'em_andamento' conforme o default do modelo
    viagem.data_inicio = datetime.utcnow()

    db.session.commit()

    return jsonify({'success': True})

@app.route('/viagens_pendentes', methods=['GET'])
@login_required
def viagens_pendentes():
    if current_user.role != 'Motorista':
        return jsonify({'success': False, 'message': 'Acesso restrito a motoristas.'}), 403

    try:
        viagens = Viagem.query.filter_by(status='pendente').all()
        viagens_data = []
        for v in viagens:
            motorista_nome = 'N/A'
            if v.motorista_id:
                motorista_nome = v.motorista_formal.nome if v.motorista_formal else 'N/A'
            elif v.motorista_cpf_cnpj:
                usuario_com_cpf = Usuario.query.filter_by(cpf_cnpj=v.motorista_cpf_cnpj).first()
                if usuario_com_cpf:
                    motorista_nome = f"{usuario_com_cpf.nome} {usuario_com_cpf.sobrenome}"
                else:
                    motorista_formal_cpf = Motorista.query.filter_by(cpf_cnpj=v.motorista_cpf_cnpj).first()
                    if motorista_formal_cpf:
                        motorista_nome = motorista_formal_cpf.nome

            viagens_data.append({
                'id': v.id,
                'cliente': v.cliente,
                'endereco_saida': v.endereco_saida,
                'endereco_destino': v.endereco_destino,
                'data_inicio': v.data_inicio.strftime('%d/%m/%Y %H:%M'),
                'distancia_km': v.distancia_km,
                'motorista_nome': motorista_nome, # Inclui o nome do motorista
                'veiculo_placa': v.veiculo.placa, # Inclui a placa do veículo
                'veiculo_modelo': v.veiculo.modelo # Inclui o modelo do veículo
            })
        return jsonify({'success': True, 'viagens': viagens_data})
    except Exception as e:
        logger.error(f"Erro ao obter viagens pendentes: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500
    
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def seed_database(force=False):
    try:
        if force:
            logger.info("Forçando limpeza do banco de dados...")
            db.drop_all()
            db.create_all()
            logger.info("Banco de dados limpo e recriado.")

        inspector = db.inspect(db.engine)
        required_tables = ['usuario', 'empresa', 'cliente', 'motorista', 'veiculo', 'viagem', 'destino', 'custo_viagem', 'localizacao', 'convite']
        existing_tables = inspector.get_table_names()
        for table in required_tables:
            if table not in existing_tables:
                logger.error(f"Tabela {table} não encontrada. Execute 'flask db upgrade'.")
                return

        if not force and Usuario.query.count() > 0:
            logger.info("Banco de dados já contém dados. Use force=True para sobrescrever.")
            return

        logger.info("Iniciando semeação completa do banco de dados...")

        # 1. Criar Empresa de Exemplo
        empresa_exemplo = Empresa(
            razao_social="TrackGo Logistica LTDA",
            nome_fantasia="TrackGo",
            cnpj="11222333000144",
            inscricao_estadual="123456789",
            endereco="Rua da Tecnologia, 123",
            cidade="Curitiba",
            estado="PR",
            cep="80000100",
            telefone="41999998888",
            email_contato="contato@trackgo.com"
        )
        db.session.add(empresa_exemplo)
        db.session.commit()
        logger.info("Empresa de exemplo criada.")

        # 2. Criar Usuários e VINCULAR à Empresa
        admin = Usuario(
            nome="João",
            sobrenome="Admin",
            email="admin@trackgo.com",
            role="Admin",
            is_admin=True,
            telefone="11987654321",
            cpf_cnpj="00000000000",
            empresa_id=empresa_exemplo.id
        )
        admin.set_password("admin123")

        master = Usuario(
            nome="Maria",
            sobrenome="Master",
            email="master@trackgo.com",
            role="Master",
            telefone="11987654322",
            cpf_cnpj="11111111111",
            empresa_id=empresa_exemplo.id
        )
        master.set_password("master123")

        motorista1 = Usuario(
            nome="Carlos",
            sobrenome="Silva",
            email="carlos@trackgo.com",
            role="Motorista",
            telefone="11987654323",
            cpf_cnpj="12345678901",
            empresa_id=empresa_exemplo.id
        )
        motorista1.set_password("motorista123")

        motorista2 = Usuario(
            nome="Ana",
            sobrenome="Souza",
            email="ana@trackgo.com",
            role="Motorista",
            telefone="21987654321",
            cpf_cnpj="98765432109",
            empresa_id=empresa_exemplo.id
        )
        motorista2.set_password("motorista123")
        
        db.session.add_all([admin, master, motorista1, motorista2])
        db.session.commit()
        logger.info("Usuários criados e vinculados à empresa.")

        # 3. Criar Clientes de Exemplo
        cliente_exemplo_1 = Cliente(
            pessoa_tipo="juridica",
            nome_razao_social="Indústrias ACME S.A.",
            nome_fantasia="ACME",
            cpf_cnpj="99888777000166",
            inscricao_estadual="ISENTO",
            cep="80230010",
            logradouro="Avenida Sete de Setembro",
            numero="3000",
            bairro="Centro",
            cidade="Curitiba",
            estado="PR",
            email="compras@acme.com",
            telefone="4133221100",
            cadastrado_por_id=admin.id
        )

        cliente_exemplo_2 = Cliente(
            pessoa_tipo="fisica",
            nome_razao_social="Maria Joaquina de Amaral Pereira",
            cpf_cnpj="11122233344",
            cep="80420000",
            logradouro="Rua Comendador Araújo",
            numero="500",
            bairro="Batel",
            cidade="Curitiba",
            estado="PR",
            email="maria.joaquina@email.com",
            telefone="41988776655",
            cadastrado_por_id=admin.id
        )
        db.session.add_all([cliente_exemplo_1, cliente_exemplo_2])
        db.session.commit()
        logger.info("Clientes de exemplo criados.")

        # 4. Criar Motoristas formais (vinculados aos usuários)
        motorista1_db = Motorista(
            nome="Carlos Silva",
            data_nascimento=datetime.strptime("1985-05-15", '%Y-%m-%d').date(),
            endereco="Rua das Flores, 123, São Paulo, SP",
            pessoa_tipo="fisica",
            cpf_cnpj="12345678901",
            rg="123456789",
            telefone="11987654323",
            cnh="98765432101",
            validade_cnh=datetime.strptime("2026-12-31", '%Y-%m-%d').date(),
            usuario_id=motorista1.id
        )

        motorista2_db = Motorista(
            nome="Ana Souza",
            data_nascimento=datetime.strptime("1990-03-22", '%Y-%m-%d').date(),
            endereco="Avenida Central, 456, Rio de Janeiro, RJ",
            pessoa_tipo="fisica",
            cpf_cnpj="98765432109",
            rg="987654321",
            telefone="21987654321",
            cnh="12345678902",
            validade_cnh=datetime.strptime("2025-10-15", '%Y-%m-%d').date(),
            usuario_id=motorista2.id
        )
        db.session.add_all([motorista1_db, motorista2_db])
        db.session.commit()
        logger.info("Motoristas formais criados.")

        # 5. Criar Veículos
        veiculo1 = Veiculo(placa="ABC1234", categoria="Caminhão", modelo="Volvo FH", ano=2020)
        veiculo2 = Veiculo(placa="XYZ5678", categoria="Van", modelo="Mercedes Sprinter", ano=2018)
        veiculo3 = Veiculo(placa="DEF9012", categoria="Carro", modelo="Fiat Toro", ano=2022, disponivel=False)
        db.session.add_all([veiculo1, veiculo2, veiculo3])
        db.session.commit()
        logger.info("Veículos criados.")

        # 6. Criar Viagens
        viagem1 = Viagem(
            motorista_id=motorista1_db.id,
            motorista_cpf_cnpj=motorista1.cpf_cnpj,
            veiculo_id=veiculo1.id,
            cliente=cliente_exemplo_1.nome_razao_social, # Usando cliente semeado
            endereco_saida="Rua das Palmeiras, 100, São Paulo, SP",
            endereco_destino="Avenida Paulista, 2000, São Paulo, SP",
            distancia_km=10.5,
            data_inicio=datetime.strptime("2025-06-01 08:00", '%Y-%m-%d %H:%M'),
            data_fim=datetime.strptime("2025-06-01 12:00", '%Y-%m-%d %H:%M'),
            duracao_segundos=14400,
            custo=150.00,
            valor_recebido=300.00,
            forma_pagamento="Pix",
            status="concluida"
        )
        viagem2 = Viagem(
            motorista_id=motorista2_db.id,
            motorista_cpf_cnpj=motorista2.cpf_cnpj,
            veiculo_id=veiculo2.id,
            cliente=cliente_exemplo_2.nome_razao_social, # Usando cliente semeado
            endereco_saida="Rua do Comércio, 50, Rio de Janeiro, RJ",
            endereco_destino="Copacabana, 300, Rio de Janeiro, RJ",
            distancia_km=15.0,
            data_inicio=datetime.strptime("2025-06-05 09:00", '%Y-%m-%d %H:%M'),
            duracao_segundos=7200,
            status="em_andamento",
            forma_pagamento="Cartão"
        )
        viagem3 = Viagem(
            motorista_cpf_cnpj=motorista1.cpf_cnpj,
            veiculo_id=veiculo3.id,
            cliente="Cliente Avulso C",
            endereco_saida="Rua Central, 200, Belo Horizonte, MG",
            endereco_destino="Praça da Liberdade, 500, Belo Horizonte, MG",
            distancia_km=8.0,
            data_inicio=datetime.strptime("2025-06-10 10:00", '%Y-%m-%d %H:%M'),
            duracao_segundos=3600,
            status="pendente",
            forma_pagamento="Boleto"
        )
        db.session.add_all([viagem1, viagem2, viagem3])
        db.session.commit()
        logger.info("Viagens criadas.")

        # 7. Criar Destinos
        db.session.add_all([
            Destino(viagem_id=viagem1.id, endereco="Avenida Paulista, 2000, São Paulo, SP", ordem=1),
            Destino(viagem_id=viagem2.id, endereco="Copacabana, 300, Rio de Janeiro, RJ", ordem=1),
            Destino(viagem_id=viagem3.id, endereco="Praça da Liberdade, 500, Belo Horizonte, MG", ordem=1)
        ])
        db.session.commit()
        logger.info("Destinos criados.")

        # 8. Criar Custos
        db.session.add(CustoViagem(viagem_id=viagem1.id, combustivel=100.00, pedagios=30.00, alimentacao=20.00))
        db.session.commit()
        logger.info("Custos criados.")

        # 9. Criar Localizações
        localizacoes_data = [
            {'motorista_id': motorista1_db.id, 'viagem_id': viagem1.id, 'latitude': -23.5505, 'longitude': -46.6333, 'timestamp': datetime.strptime("2025-06-01 09:00", '%Y-%m-%d %H:%M')},
            {'motorista_id': motorista2_db.id, 'viagem_id': viagem2.id, 'latitude': -22.9068, 'longitude': -43.1729, 'timestamp': datetime.strptime("2025-06-05 10:00", '%Y-%m-%d %H:%M')}
        ]
        for loc_data in localizacoes_data:
            endereco = get_address_geoapify(loc_data['latitude'], loc_data['longitude'])
            db.session.add(Localizacao(
                motorista_id=loc_data['motorista_id'],
                viagem_id=loc_data['viagem_id'],
                latitude=loc_data['latitude'],
                longitude=loc_data['longitude'],
                endereco=endereco,
                timestamp=loc_data['timestamp']
            ))
        db.session.commit()
        logger.info("Localizações criadas.")
        logger.info("Semeação do banco de dados concluída com sucesso!")

    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao semear o banco de dados: {str(e)}", exc_info=True)
        raise

def get_address_geoapify(lat, lon):
    try:
        url = f'https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey={GEOAPIFY_API_KEY}'
        response = requests.get(url)
        response.raise_for_status() 
        data = response.json()
        if data['features']:
            return data['features'][0]['properties']['formatted']
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro de rede/API na geocodificação Geoapify: {str(e)}", exc_info=True)
    except Exception as e:
        logger.error(f"Erro inesperado na geocodificação Geoapify: {str(e)}", exc_info=True)
    return "Endereço não encontrado"



@app.route('/ultima_localizacao/<int:viagem_id>', methods=['GET'])
def ultima_localizacao(viagem_id):
    """Retorna a última localização registrada para uma viagem, incluindo coordenadas."""
    localizacao = Localizacao.query.filter_by(viagem_id=viagem_id).order_by(Localizacao.timestamp.desc()).first()
    
    if localizacao:
        return jsonify({
            'success': True, 
            'endereco': localizacao.endereco,
            'latitude': localizacao.latitude,    # <-- LINHA ADICIONADA
            'longitude': localizacao.longitude   # <-- LINHA ADICIONADA
        })
    
    return jsonify({'success': False, 'message': 'Nenhuma localização encontrada para esta viagem.'})


@app.route('/motorista/<int:motorista_id>/perfil', endpoint='perfil_motorista')
@login_required 
def perfil_motorista(motorista_id):
    
    motorista = Motorista.query.get_or_404(motorista_id)

    viagens = Viagem.query.filter(
        or_(
            Viagem.motorista_id == motorista.id,
            Viagem.motorista_cpf_cnpj == motorista.cpf_cnpj
        )
    ).order_by(Viagem.data_inicio.desc()).all() 

    # 3. Calcula estatísticas para exibir no perfil (opcional, mas muito útil)
    total_viagens = len(viagens)
    total_distancia = sum(v.distancia_km or 0 for v in viagens)
    total_receita = sum(v.valor_recebido or 0 for v in viagens)
    total_custo = sum(v.custo or 0 for v in viagens)
    lucro_total = total_receita - total_custo

    stats = {
        'total_viagens': total_viagens,
        'total_distancia': round(total_distancia, 2),
        'total_receita': round(total_receita, 2),
        'total_custo': round(total_custo, 2),
        'lucro_total': round(lucro_total, 2)
    }

    # 4. Renderiza um novo template HTML, passando os dados do motorista e suas viagens.
    return render_template('perfil_motorista.html', motorista=motorista, viagens=viagens, stats=stats)

# Adicione esta nova rota em app.py

@app.route('/romaneio/viagem/<int:viagem_id>', methods=['GET', 'POST'])
@login_required
def gerar_romaneio(viagem_id):
    viagem = Viagem.query.get_or_404(viagem_id)
    romaneio = Romaneio.query.filter_by(viagem_id=viagem_id).first()

    if request.method == 'POST':
        # Esta seção agora lida com CRIAR e ATUALIZAR
        data_emissao_str = request.form.get('data_emissao')
        observacoes = request.form.get('observacoes')
        data_emissao = datetime.strptime(data_emissao_str, '%Y-%m-%d').date() if data_emissao_str else datetime.utcnow().date()

        if romaneio:
            # --- LÓGICA DE ATUALIZAÇÃO ---
            romaneio.data_emissao = data_emissao
            romaneio.observacoes = observacoes
            
            # Remove itens antigos para adicionar os novos (maneira mais simples de sincronizar)
            ItemRomaneio.query.filter_by(romaneio_id=romaneio.id).delete()
            flash_message = 'Romaneio atualizado com sucesso!'
        else:
            # --- LÓGICA DE CRIAÇÃO ---
            romaneio = Romaneio(
                viagem_id=viagem.id,
                data_emissao=data_emissao,
                observacoes=observacoes
            )
            db.session.add(romaneio)
            db.session.flush() # Necessário para obter o romaneio.id para os itens
            flash_message = 'Romaneio salvo com sucesso!'

        # Processar/Reprocessar os itens da carga
        item_counter = 1
        while f'produto_{item_counter}' in request.form:
            produto = request.form.get(f'produto_{item_counter}')
            if produto: # Processa apenas se o campo produto não estiver vazio
                item = ItemRomaneio(
                    romaneio_id=romaneio.id,
                    produto_descricao=produto,
                    quantidade=int(request.form.get(f'qtd_{item_counter}', 1)),
                    embalagem=request.form.get(f'embalagem_{item_counter}'),
                    peso_kg=float(request.form.get(f'peso_{item_counter}', 0))
                )
                db.session.add(item)
            item_counter += 1
            
        db.session.commit()
        flash(flash_message, 'success')
        
        # Padrão Post/Redirect/Get: Redireciona para a mesma página com o método GET
        return redirect(url_for('gerar_romaneio', viagem_id=viagem_id))

    # --- LÓGICA GET (Carregar a página) ---
    if romaneio:
        # Se um romaneio já existe, passa o objeto para o template
        return render_template('cadastro_romaneio.html', viagem=viagem, romaneio=romaneio)
    else:
        # Se não existe, prepara os dados para um novo romaneio
        motorista_nome = 'N/A'
        if viagem.motorista_id:
            motorista_nome = viagem.motorista_formal.nome if viagem.motorista_formal else 'N/A'
        elif viagem.motorista_cpf_cnpj:
            usuario = Usuario.query.filter_by(cpf_cnpj=viagem.motorista_cpf_cnpj).first()
            if usuario: motorista_nome = f"{usuario.nome} {usuario.sobrenome}"
        
        dados_novo_romaneio = {
            'dest_nome': viagem.cliente,
            'dest_endereco': viagem.endereco_destino,
            'transportadora': motorista_nome,
            'placa_veiculo': viagem.veiculo.placa
        }
        # Gera um número de romaneio sequencial (pode ser o próximo ID)
        ultimo_id = db.session.query(db.func.max(Romaneio.id)).scalar() or 0
        
        return render_template('cadastro_romaneio.html', 
                               viagem=viagem, 
                               romaneio=None, # Importante: envia None para indicar que é novo
                               dados=dados_novo_romaneio,
                               numero_romaneio=ultimo_id + 1)
    
@app.route('/consultar_romaneios')
@login_required
def consultar_romaneios():
    search_query = request.args.get('search', '').strip()
    query = Romaneio.query.join(Viagem)  # Join com Viagem para filtros adicionais
    
    if search_query:
        query = query.filter(
            or_(
                Viagem.cliente.ilike(f'%{search_query}%'),
                Viagem.motorista_formal.has(Motorista.nome.ilike(f'%{search_query}%')),
                Viagem.veiculo.has(Veiculo.placa.ilike(f'%{search_query}%'))
            )
        )
    
    romaneios = query.order_by(Romaneio.data_emissao.desc()).all()
    
    return render_template(
        'consultar_romaneios.html',
        romaneios=romaneios,
        search_query=search_query,
        active_page='consultar_romaneios'
    )

@app.template_filter('get_usuario')
def get_usuario(cpf_cnpj):
    return Usuario.query.filter_by(cpf_cnpj=cpf_cnpj).first()

@app.route('/romaneio/<int:romaneio_id>', methods=['GET'])
@login_required
def visualizar_romaneio(romaneio_id):
    romaneio = Romaneio.query.get_or_404(romaneio_id)
    return render_template('visualizar_romaneio.html', romaneio=romaneio, active_page='consultar_romaneios')

import click

@app.cli.command("create-owner")
@click.argument("email")
@click.argument("password")
def create_owner_command(email, password):
    """Cria um novo usuário com o papel de Owner."""
    
    if Usuario.query.filter_by(email=email).first():
        print(f"Erro: O usuário com o e-mail '{email}' já existe.")
        return

    try:
        owner = Usuario(
            nome='Proprietário',
            sobrenome='do Sistema',
            email=email,
            role='Owner',
            is_admin=True # Um Owner também pode ser admin
        )
        owner.set_password(password)
        db.session.add(owner)
        db.session.commit()
        print(f"Usuário Owner '{email}' criado com sucesso!")
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao criar o usuário Owner: {e}")

@socketio.on('connect')
def handle_connect():
    logger.info(f"Cliente conectado: {request.sid}")

@socketio.on('disconnect')
def handle_disconnect():
    logger.info(f"Cliente desconectado: {request.sid}")

@socketio.on('join_trip_room')
def handle_join_trip_room(data):
    viagem_id = data.get('viagem_id')
    if viagem_id:
        room = f'viagem_{viagem_id}'
        join_room(room)
        logger.info(f"Cliente {request.sid} entrou na sala da viagem {viagem_id}")

@socketio.on('leave_trip_room')
def handle_leave_trip_room(data):
    viagem_id = data.get('viagem_id')
    if viagem_id:
        room = f'viagem_{viagem_id}'
        leave_room(room)
        logger.info(f"Cliente {request.sid} saiu da sala da viagem {viagem_id}")

@socketio.on('atualizar_localizacao_socket')
@login_required
def handle_location_update(data):
    # VERSÃO ÚNICA E CORRETA DESTA FUNÇÃO
    logger.info(f"--- PONTO 1: handle_location_update chamada com dados: {data}")
    lat, lon, viagem_id = data.get('latitude'), data.get('longitude'), data.get('viagem_id')

    if not all([lat, lon, viagem_id]): return

    try:
        motorista_formal = Motorista.query.filter_by(cpf_cnpj=current_user.cpf_cnpj).first()
        if not motorista_formal:
            logger.error(f"--- PONTO 2 FALHA: Motorista com CPF {current_user.cpf_cnpj} não encontrado.")
            return
        
        logger.info(f"--- PONTO 2 SUCESSO: Motorista encontrado: {motorista_formal.nome}")

        endereco = get_address_geoapify(lat, lon)
        nova_localizacao = Localizacao(motorista_id=motorista_formal.id, viagem_id=viagem_id, latitude=lat, longitude=lon, endereco=endereco)
        db.session.add(nova_localizacao)
        db.session.commit()
        
        room = f'viagem_{viagem_id}'
        payload = {'latitude': lat, 'longitude': lon, 'endereco': endereco, 'viagem_id': viagem_id}
        
        logger.info(f"--- PONTO 3: Emitindo para a sala '{room}'")
        socketio.emit('localizacao_atualizada', payload, to=room)
        logger.info("--- PONTO 4: Emissão concluída.")
    except Exception as e:
        logger.error(f"Erro no evento de localização: {e}", exc_info=True)

@app.route('/manifest.json')
def manifest():
    return send_from_directory('templates', 'manifest.json')

# Rota para servir o sw.js a partir da raiz do projeto
@app.route('/sw.js')
def service_worker():
    response = make_response(send_from_directory('.', 'sw.js'))
    response.headers['Content-Type'] = 'application/javascript'
    return response

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        seed_database(force=False)
    socketio.run(app, debug=True)
